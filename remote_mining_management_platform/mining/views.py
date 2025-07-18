from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.urls import reverse_lazy
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.http import HttpResponse
from openpyxl import Workbook
from .models import RemoteMiningPlatform, Miner, Settings, APIData, Payout
from .forms import RemoteMiningPlatformForm, MinerForm, SettingsForm, PayoutForm
from .api_utils import fetch_all_api_data


class PlatformListView(ListView):
    model = RemoteMiningPlatform
    template_name = 'mining/platform_list.html'
    context_object_name = 'platforms'
    paginate_by = 10


class PlatformDetailView(DetailView):
    model = RemoteMiningPlatform
    template_name = 'mining/platform_detail.html'
    context_object_name = 'platform'


class PlatformCreateView(CreateView):
    model = RemoteMiningPlatform
    form_class = RemoteMiningPlatformForm
    template_name = 'mining/platform_form.html'
    success_url = reverse_lazy('platform_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Platform created successfully.')
        return super().form_valid(form)


class PlatformUpdateView(UpdateView):
    model = RemoteMiningPlatform
    form_class = RemoteMiningPlatformForm
    template_name = 'mining/platform_form.html'
    success_url = reverse_lazy('platform_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Platform updated successfully.')
        return super().form_valid(form)


class PlatformDeleteView(DeleteView):
    model = RemoteMiningPlatform
    template_name = 'mining/platform_confirm_delete.html'
    success_url = reverse_lazy('platform_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, "Platform deleted successfully!")
        return super().delete(request, *args, **kwargs)


# Miner Views
class MinerListView(ListView):
    model = Miner
    template_name = 'mining/miner_list.html'
    context_object_name = 'miners'
    paginate_by = 12


class MinerDetailView(DetailView):
    model = Miner
    template_name = 'mining/miner_detail.html'
    context_object_name = 'miner'


class MinerCreateView(CreateView):
    model = Miner
    form_class = MinerForm
    template_name = 'mining/miner_form.html'
    success_url = reverse_lazy('miner_list')

    def form_valid(self, form):
        messages.success(self.request, "Miner created successfully!")
        return super().form_valid(form)


class MinerUpdateView(UpdateView):
    model = Miner
    form_class = MinerForm
    template_name = 'mining/miner_form.html'
    success_url = reverse_lazy('miner_list')

    def form_valid(self, form):
        messages.success(self.request, "Miner updated successfully!")
        return super().form_valid(form)


class MinerDeleteView(DeleteView):
    model = Miner
    template_name = 'mining/miner_confirm_delete.html'
    success_url = reverse_lazy('miner_list')
    context_object_name = 'miner'

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Miner deleted successfully!")
        return super().delete(request, *args, **kwargs)


# Payout Views
class PayoutListView(ListView):
    model = Payout
    template_name = 'mining/payout_list.html'
    context_object_name = 'payouts'
    paginate_by = 10


class PayoutDetailView(DetailView):
    model = Payout
    template_name = 'mining/payout_detail.html'
    context_object_name = 'payout'


class PayoutCreateView(CreateView):
    model = Payout
    form_class = PayoutForm
    template_name = 'mining/payout_form.html'
    success_url = reverse_lazy('payout_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Payout added successfully!')
        return super().form_valid(form)


class PayoutUpdateView(UpdateView):
    model = Payout
    form_class = PayoutForm
    template_name = 'mining/payout_form.html'
    success_url = reverse_lazy('payout_list')
    
    def form_valid(self, form):
        messages.success(self.request, 'Payout updated successfully!')
        return super().form_valid(form)


class PayoutDeleteView(DeleteView):
    model = Payout
    template_name = 'mining/payout_confirm_delete.html'
    success_url = reverse_lazy('payout_list')
    context_object_name = 'payout'
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Payout deleted successfully!')
        return super().delete(request, *args, **kwargs)


def api_data_view(request):
    """API Data page view"""
    api_data = APIData.get_api_data()
    
    if request.method == 'POST':
        # Fetch API data when button is clicked
        result = fetch_all_api_data()
        
        if result['success']:
            # Update the API data in database
            api_data.bitcoin_price_usd = result['bitcoin_price_usd']
            api_data.network_hashrate_ehs = result['network_hashrate_ehs']
            api_data.network_difficulty = result['network_difficulty']
            api_data.save()
            
            messages.success(request, result['message'])
        else:
            messages.error(request, result['message'])
            
        return redirect('api_data')
    
    return render(request, 'mining/api_data.html', {'api_data': api_data})


def settings_view(request):
    """Settings page view"""
    settings = Settings.get_settings()
    
    if request.method == 'POST':
        form = SettingsForm(request.POST, instance=settings)
        if form.is_valid():
            form.save()
            messages.success(request, 'Settings saved successfully!')
            return redirect('settings')
    else:
        form = SettingsForm(instance=settings)
    
    return render(request, 'mining/settings.html', {'form': form})


# Import Template Download Views
def download_platform_template(request):
    """Download import template for Remote Mining Platforms"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Platform Import Template"
    
    # Add headers based on form fields
    headers = ['name', 'website_link', 'portal_url', 'point_of_contact_name', 
               'point_of_contact_email', 'point_of_contact_phone', 'point_of_contact_telegram', 'energy_price']
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="platform_import_template.xlsx"'
    wb.save(response)
    return response


def download_miner_template(request):
    """Download import template for Miners"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Miner Import Template"
    
    # Add headers based on form fields (excluding image field for import)
    headers = ['model', 'manufacturer', 'product_link', 'serial_number', 
               'platform', 'platform_internal_id', 'hashrate', 'power', 'efficiency', 
               'purchase_price', 'purchase_date', 'start_date', 'location']
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="miner_import_template.xlsx"'
    wb.save(response)
    return response


def download_payout_template(request):
    """Download import template for Payouts"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Payout Import Template"
    
    # Add headers based on form fields
    headers = ['payout_date', 'payout_amount', 'platform', 'transaction_id']
    
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="payout_import_template.xlsx"'
    wb.save(response)
    return response
