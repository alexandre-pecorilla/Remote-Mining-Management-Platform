from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.urls import reverse_lazy, reverse
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.views.decorators.http import require_POST
from django.db.models import Sum, Q
from django.db.models.functions import TruncMonth
from openpyxl import Workbook, load_workbook
from decimal import Decimal, InvalidOperation
from datetime import datetime, date, timedelta
import json
import logging
import threading
import time

logger = logging.getLogger(__name__)
from .models import RemoteMiningPlatform, Miner, Settings, APIData, Payout, Expense, TopUp
from .forms import RemoteMiningPlatformForm, MinerForm, SettingsForm, PayoutForm, ExpenseForm, TopUpForm
from .api_utils import fetch_all_api_data, get_historical_btc_price
from .services import get_capex_opex_data, get_income_data, get_overview_data, get_forecasting_data, resolve_selected_platform


# Home Page View
def home_view(request):
    """Home page with navigation to all sections of the application"""
    return render(request, 'mining/home.html')


# CAPEX/OPEX Dashboard View
def capex_opex_dashboard(request):
    """Dashboard view for CAPEX/OPEX analysis"""
    data = get_capex_opex_data()
    return render(request, 'mining/capex_opex_dashboard.html', data)


def export_capex_opex_data(request):
    """Export CAPEX/OPEX dashboard data to Excel file"""
    
    wb = Workbook()
    data = get_capex_opex_data()
    total_expenses = data['total_expenses']
    total_capex = data['total_capex']
    total_opex = data['total_opex']
    platform_expenses = data['platform_expenses']
    monthly_capex = data['monthly_capex']
    monthly_capex_by_platform = data['monthly_capex_by_platform']
    monthly_opex = data['monthly_opex']
    monthly_opex_by_platform = data['monthly_opex_by_platform']
    all_months = data['all_months']

    # Sheet 1: Total Expenses Summary
    ws_summary = wb.create_sheet(title='Total Expenses Summary')
    
    # Headers
    ws_summary.cell(row=1, column=1, value='Expense Type')
    ws_summary.cell(row=1, column=2, value='Amount (USD)')
    
    # Data rows
    ws_summary.cell(row=2, column=1, value='Total Expenses')
    ws_summary.cell(row=2, column=2, value=float(total_expenses))
    
    ws_summary.cell(row=3, column=1, value='Total CAPEX')
    ws_summary.cell(row=3, column=2, value=float(total_capex))
    
    ws_summary.cell(row=4, column=1, value='Total OPEX')
    ws_summary.cell(row=4, column=2, value=float(total_opex))
    
    # Sheet 2: Expenses by Platform
    if platform_expenses:
        ws_platform = wb.create_sheet(title='Expenses by Platform')
        
        # Headers
        ws_platform.cell(row=1, column=1, value='Platform')
        ws_platform.cell(row=1, column=2, value='Total Expenses (USD)')
        ws_platform.cell(row=1, column=3, value='CAPEX (USD)')
        ws_platform.cell(row=1, column=4, value='OPEX (USD)')
        
        # Data rows
        for row, item in enumerate(platform_expenses, start=1):
            ws_platform.cell(row=row + 1, column=1, value=item['platform'].name)
            ws_platform.cell(row=row + 1, column=2, value=float(item['total']))
            ws_platform.cell(row=row + 1, column=3, value=float(item['capex']))
            ws_platform.cell(row=row + 1, column=4, value=float(item['opex']))
    
    # Sheet 3: Monthly CAPEX
    if monthly_capex and all_months:
        ws_monthly_capex = wb.create_sheet(title='Monthly CAPEX')
        
        # Headers
        ws_monthly_capex.cell(row=1, column=1, value='Month')
        ws_monthly_capex.cell(row=1, column=2, value='Total CAPEX (USD)')
        
        # Platform headers
        col = 2
        platform_cols = {}
        for platform in monthly_capex_by_platform.keys():
            ws_monthly_capex.cell(row=1, column=col + 1, value=f'{platform.name} CAPEX (USD)')
            platform_cols[platform] = col
            col += 1
        
        # Data rows
        for row, month in enumerate(all_months, start=1):
            if month:
                ws_monthly_capex.cell(row=row + 1, column=1, value=month.strftime('%Y-%m'))
                
                # Total CAPEX for this month
                month_total = Decimal('0')
                for item in monthly_capex:
                    if item['month'] == month:
                        month_total = item['total']
                        break
                ws_monthly_capex.cell(row=row + 1, column=2, value=float(month_total))
                
                # Platform CAPEX for this month
                for platform, platform_data in monthly_capex_by_platform.items():
                    platform_month_total = Decimal('0')
                    for item in platform_data:
                        if item['month'] == month:
                            platform_month_total = item['total']
                            break
                    ws_monthly_capex.cell(row=row + 1, column=platform_cols[platform] + 1, value=float(platform_month_total))
    
    # Sheet 4: Monthly OPEX
    if monthly_opex and all_months:
        ws_monthly_opex = wb.create_sheet(title='Monthly OPEX')
        
        # Headers
        ws_monthly_opex.cell(row=1, column=1, value='Month')
        ws_monthly_opex.cell(row=1, column=2, value='Total OPEX (USD)')
        
        # Platform headers
        col = 2
        platform_cols = {}
        for platform in monthly_opex_by_platform.keys():
            ws_monthly_opex.cell(row=1, column=col + 1, value=f'{platform.name} OPEX (USD)')
            platform_cols[platform] = col
            col += 1
        
        # Data rows
        for row, month in enumerate(all_months, start=1):
            if month:
                ws_monthly_opex.cell(row=row + 1, column=1, value=month.strftime('%Y-%m'))
                
                # Total OPEX for this month
                month_total = Decimal('0')
                for item in monthly_opex:
                    if item['month'] == month:
                        month_total = item['total']
                        break
                ws_monthly_opex.cell(row=row + 1, column=2, value=float(month_total))
                
                # Platform OPEX for this month
                for platform, platform_data in monthly_opex_by_platform.items():
                    platform_month_total = Decimal('0')
                    for item in platform_data:
                        if item['month'] == month:
                            platform_month_total = item['total']
                            break
                    ws_monthly_opex.cell(row=row + 1, column=platform_cols[platform] + 1, value=float(platform_month_total))
    
    # Generate response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="capex_opex_dashboard_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response


# Income Dashboard View
def income_dashboard(request):
    """Dashboard view for Income analysis"""
    data = get_income_data()
    current_btc_price = data['current_btc_price']
    platform_income = data['platform_income']
    monthly_income_btc = data['monthly_income_btc']
    monthly_income_by_platform = data['monthly_income_by_platform']
    all_months = data['all_months']

    # Prepare monthly data for template (dashboard-specific pivot)
    monthly_btc_data = []
    monthly_usd_then_data = []
    monthly_usd_now_data = []
    
    for month in all_months:
        if month:
            # BTC data
            btc_row = {'month': month, 'total': Decimal('0'), 'platforms': {}}
            usd_then_row = {'month': month, 'total': Decimal('0'), 'platforms': {}}
            usd_now_row = {'month': month, 'total': Decimal('0'), 'platforms': {}}
            
            # Find total for this month
            for item in monthly_income_btc:
                if item['month'] == month:
                    btc_row['total'] = item['total_btc']
                    usd_then_row['total'] = item['total_usd_then']
                    usd_now_row['total'] = item['total_usd_now']
                    break
            
            # Add platform data
            for platform, platform_data in monthly_income_by_platform.items():
                btc_row['platforms'][platform] = Decimal('0')
                usd_then_row['platforms'][platform] = Decimal('0')
                usd_now_row['platforms'][platform] = Decimal('0')
                
                for item in platform_data:
                    if item['month'] == month:
                        btc_row['platforms'][platform] = item['total_btc']
                        usd_then_row['platforms'][platform] = item['total_usd_then']
                        usd_now_row['platforms'][platform] = item['total_usd_now']
                        break
            
            monthly_btc_data.append(btc_row)
            monthly_usd_then_data.append(usd_then_row)
            monthly_usd_now_data.append(usd_now_row)
    
    context = {
        'total_income_btc': data['total_income_btc'],
        'total_income_usd_then': data['total_income_usd_then'],
        'total_income_usd_now': data['total_income_usd_now'],
        'platform_income': platform_income,
        'monthly_btc_data': monthly_btc_data,
        'monthly_usd_then_data': monthly_usd_then_data,
        'monthly_usd_now_data': monthly_usd_now_data,
        'platforms_with_income': [item['platform'] for item in platform_income],
        'current_btc_price': current_btc_price,
    }

    return render(request, 'mining/income_dashboard.html', context)


def export_income_data(request):
    """Export Income dashboard data to Excel file"""
    
    wb = Workbook()
    data = get_income_data()
    current_btc_price = data['current_btc_price']
    total_income_btc = data['total_income_btc']
    total_income_usd_then = data['total_income_usd_then']
    total_income_usd_now = data['total_income_usd_now']
    platform_income = data['platform_income']
    monthly_income_btc = data['monthly_income_btc']
    monthly_income_by_platform = data['monthly_income_by_platform']
    all_months = data['all_months']

    # Sheet 1: Total Income Summary
    ws_summary = wb.create_sheet(title='Total Income Summary')
    
    # Headers
    ws_summary.cell(row=1, column=1, value='Income Type')
    ws_summary.cell(row=1, column=2, value='Amount')
    
    # Data rows
    ws_summary.cell(row=2, column=1, value='Total Income BTC')
    ws_summary.cell(row=2, column=2, value=float(total_income_btc))
    
    ws_summary.cell(row=3, column=1, value='Total Income USD (then)')
    ws_summary.cell(row=3, column=2, value=float(total_income_usd_then))
    
    ws_summary.cell(row=4, column=1, value='Total Income USD (now)')
    ws_summary.cell(row=4, column=2, value=float(total_income_usd_now))
    
    # Sheet 2: Income by Platform
    if platform_income:
        ws_platform = wb.create_sheet(title='Income by Platform')
        
        # Headers
        ws_platform.cell(row=1, column=1, value='Platform')
        ws_platform.cell(row=1, column=2, value='Total Income BTC')
        ws_platform.cell(row=1, column=3, value='Total Income USD (then)')
        ws_platform.cell(row=1, column=4, value='Total Income USD (now)')
        
        # Data rows
        for row, item in enumerate(platform_income, start=1):
            ws_platform.cell(row=row + 1, column=1, value=item['platform'].name)
            ws_platform.cell(row=row + 1, column=2, value=float(item['total_btc']))
            ws_platform.cell(row=row + 1, column=3, value=float(item['total_usd_then']))
            ws_platform.cell(row=row + 1, column=4, value=float(item['total_usd_now']))
    
    # Sheet 3: Monthly Income BTC
    if monthly_income_btc and all_months:
        ws_monthly_btc = wb.create_sheet(title='Monthly Income BTC')
        
        # Headers
        ws_monthly_btc.cell(row=1, column=1, value='Month')
        ws_monthly_btc.cell(row=1, column=2, value='Total Income BTC')
        
        # Platform headers
        col = 2
        platform_cols = {}
        for platform in monthly_income_by_platform.keys():
            ws_monthly_btc.cell(row=1, column=col + 1, value=f'{platform.name} BTC')
            platform_cols[platform] = col
            col += 1
        
        # Data rows
        for row, month in enumerate(all_months, start=1):
            if month:
                ws_monthly_btc.cell(row=row + 1, column=1, value=month.strftime('%Y-%m'))
                
                # Total BTC for this month
                month_total = Decimal('0')
                for item in monthly_income_btc:
                    if item['month'] == month:
                        month_total = item['total_btc']
                        break
                ws_monthly_btc.cell(row=row + 1, column=2, value=float(month_total))
                
                # Platform BTC for this month
                for platform, platform_data in monthly_income_by_platform.items():
                    platform_month_total = Decimal('0')
                    for item in platform_data:
                        if item['month'] == month:
                            platform_month_total = item['total_btc']
                            break
                    ws_monthly_btc.cell(row=row + 1, column=platform_cols[platform] + 1, value=float(platform_month_total))
    
    # Sheet 4: Monthly Income USD (then)
    if monthly_income_btc and all_months:
        ws_monthly_usd_then = wb.create_sheet(title='Monthly Income USD then')
        
        # Headers
        ws_monthly_usd_then.cell(row=1, column=1, value='Month')
        ws_monthly_usd_then.cell(row=1, column=2, value='Total Income USD (then)')
        
        # Platform headers
        col = 2
        platform_cols = {}
        for platform in monthly_income_by_platform.keys():
            ws_monthly_usd_then.cell(row=1, column=col + 1, value=f'{platform.name} USD (then)')
            platform_cols[platform] = col
            col += 1
        
        # Data rows
        for row, month in enumerate(all_months, start=1):
            if month:
                ws_monthly_usd_then.cell(row=row + 1, column=1, value=month.strftime('%Y-%m'))
                
                # Total USD then for this month
                month_total = Decimal('0')
                for item in monthly_income_btc:
                    if item['month'] == month:
                        month_total = item['total_usd_then']
                        break
                ws_monthly_usd_then.cell(row=row + 1, column=2, value=float(month_total))
                
                # Platform USD then for this month
                for platform, platform_data in monthly_income_by_platform.items():
                    platform_month_total = Decimal('0')
                    for item in platform_data:
                        if item['month'] == month:
                            platform_month_total = item['total_usd_then']
                            break
                    ws_monthly_usd_then.cell(row=row + 1, column=platform_cols[platform] + 1, value=float(platform_month_total))
    
    # Sheet 5: Monthly Income USD (now)
    if monthly_income_btc and all_months:
        ws_monthly_usd_now = wb.create_sheet(title='Monthly Income USD now')
        
        # Headers
        ws_monthly_usd_now.cell(row=1, column=1, value='Month')
        ws_monthly_usd_now.cell(row=1, column=2, value='Total Income USD (now)')
        
        # Platform headers
        col = 2
        platform_cols = {}
        for platform in monthly_income_by_platform.keys():
            ws_monthly_usd_now.cell(row=1, column=col + 1, value=f'{platform.name} USD (now)')
            platform_cols[platform] = col
            col += 1
        
        # Data rows
        for row, month in enumerate(all_months, start=1):
            if month:
                ws_monthly_usd_now.cell(row=row + 1, column=1, value=month.strftime('%Y-%m'))
                
                # Total USD now for this month
                month_total = Decimal('0')
                for item in monthly_income_btc:
                    if item['month'] == month:
                        month_total = item['total_usd_now']
                        break
                ws_monthly_usd_now.cell(row=row + 1, column=2, value=float(month_total))
                
                # Platform USD now for this month
                for platform, platform_data in monthly_income_by_platform.items():
                    platform_month_total = Decimal('0')
                    for item in platform_data:
                        if item['month'] == month:
                            platform_month_total = item['total_usd_now']
                            break
                    ws_monthly_usd_now.cell(row=row + 1, column=platform_cols[platform] + 1, value=float(platform_month_total))
    
    # Generate response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="income_dashboard_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response


class PlatformListView(ListView):
    model = RemoteMiningPlatform
    template_name = 'mining/platform_list.html'
    context_object_name = 'platforms'
    paginate_by = 10


class PlatformDetailView(DetailView):
    model = RemoteMiningPlatform
    template_name = 'mining/platform_detail.html'
    context_object_name = 'platform'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_platform = self.get_object()
        
        # Navigate by name (matching list view order: alphabetical)
        # Previous = earlier in alphabet
        previous_platform = RemoteMiningPlatform.objects.filter(
            Q(name__lt=current_platform.name) |
            Q(name=current_platform.name, id__lt=current_platform.id)
        ).order_by('-name', '-id').first()
        
        # Next = later in alphabet
        next_platform = RemoteMiningPlatform.objects.filter(
            Q(name__gt=current_platform.name) |
            Q(name=current_platform.name, id__gt=current_platform.id)
        ).order_by('name', 'id').first()
        
        context['previous_platform'] = previous_platform
        context['next_platform'] = next_platform
        return context


class PlatformCreateView(CreateView):
    model = RemoteMiningPlatform
    form_class = RemoteMiningPlatformForm
    template_name = 'mining/platform_form.html'
    
    def get_success_url(self):
        return reverse_lazy('platform_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Platform created successfully.')
        return super().form_valid(form)


class PlatformUpdateView(UpdateView):
    model = RemoteMiningPlatform
    form_class = RemoteMiningPlatformForm
    template_name = 'mining/platform_form.html'
    
    def get_success_url(self):
        return reverse_lazy('platform_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Platform updated successfully.')
        return super().form_valid(form)


class PlatformDeleteView(DeleteView):
    model = RemoteMiningPlatform
    template_name = 'mining/platform_confirm_delete.html'
    success_url = reverse_lazy('platform_list')
    
    def delete(self, request, *args, **kwargs):
        messages.success(request, "Platform deleted successfully!")
        return super().delete(request, *args, **kwargs)


# Miner Views
class MinerListView(ListView):
    model = Miner
    template_name = 'mining/miner_list.html'
    context_object_name = 'miners'
    paginate_by = 50
    queryset = Miner.objects.select_related('platform')


class MinerDetailView(DetailView):
    model = Miner
    template_name = 'mining/miner_detail.html'
    context_object_name = 'miner'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_miner = self.get_object()
        
        # Get previous miner (lower ID)
        previous_miner = Miner.objects.filter(
            id__lt=current_miner.id
        ).order_by('-id').first()
        
        # Get next miner (higher ID)
        next_miner = Miner.objects.filter(
            id__gt=current_miner.id
        ).order_by('id').first()
        
        context['previous_miner'] = previous_miner
        context['next_miner'] = next_miner
        return context


class MinerCreateView(CreateView):
    model = Miner
    form_class = MinerForm
    template_name = 'mining/miner_form.html'
    
    def get_success_url(self):
        return reverse_lazy('miner_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, "Miner created successfully!")
        return super().form_valid(form)


class MinerUpdateView(UpdateView):
    model = Miner
    form_class = MinerForm
    template_name = 'mining/miner_form.html'
    
    def get_success_url(self):
        return reverse_lazy('miner_detail', kwargs={'pk': self.object.pk})

    def form_valid(self, form):
        messages.success(self.request, "Miner updated successfully!")
        return super().form_valid(form)


class MinerDeleteView(DeleteView):
    model = Miner
    template_name = 'mining/miner_confirm_delete.html'
    success_url = reverse_lazy('miner_list')
    context_object_name = 'miner'

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Miner deleted successfully!")
        return super().delete(request, *args, **kwargs)


@require_POST
def toggle_miner_active(request, pk):
    """Toggle a miner's is_active status (on/off)"""
    miner = get_object_or_404(Miner, pk=pk)
    miner.is_active = not miner.is_active
    miner.save(update_fields=['is_active'])
    status = "ON" if miner.is_active else "OFF"
    messages.success(request, f"{miner.model} turned {status}.")
    return redirect('miner_detail', pk=pk)


# Payout Views
class PayoutListView(ListView):
    model = Payout
    template_name = 'mining/payout_list.html'
    context_object_name = 'payouts'
    paginate_by = 50
    queryset = Payout.objects.select_related('platform')


class PayoutDetailView(DetailView):
    model = Payout
    template_name = 'mining/payout_detail.html'
    context_object_name = 'payout'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_payout = self.get_object()
        
        # Navigate by payout_date (matching list view order: newest first)
        # Previous = newer payout (appears before in list)
        previous_payout = Payout.objects.filter(
            Q(payout_date__gt=current_payout.payout_date) |
            Q(payout_date=current_payout.payout_date, id__gt=current_payout.id)
        ).order_by('payout_date', 'id').first()
        
        # Next = older payout (appears after in list)
        next_payout = Payout.objects.filter(
            Q(payout_date__lt=current_payout.payout_date) |
            Q(payout_date=current_payout.payout_date, id__lt=current_payout.id)
        ).order_by('-payout_date', '-id').first()
        
        context['previous_payout'] = previous_payout
        context['next_payout'] = next_payout
        return context


class PayoutCreateView(CreateView):
    model = Payout
    form_class = PayoutForm
    template_name = 'mining/payout_form.html'
    
    def get_success_url(self):
        return reverse_lazy('payout_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Payout added successfully!')
        return super().form_valid(form)


class PayoutUpdateView(UpdateView):
    model = Payout
    form_class = PayoutForm
    template_name = 'mining/payout_form.html'
    
    def get_success_url(self):
        return reverse_lazy('payout_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Payout updated successfully!')
        return super().form_valid(form)


class PayoutDeleteView(DeleteView):
    model = Payout
    template_name = 'mining/payout_confirm_delete.html'
    success_url = reverse_lazy('payout_list')
    context_object_name = 'payout'
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Payout deleted successfully!')
        return super().delete(request, *args, **kwargs)


@require_POST
def fetch_closing_price(request, payout_id):
    """Fetch historical BTC price for payout date and update closing_price field"""
    try:
        payout = get_object_or_404(Payout, pk=payout_id)

        if not payout.payout_date:
            return JsonResponse({
                'success': False,
                'error': 'Payout date is required to fetch closing price'
            })

        # Fetch historical BTC price for the payout date
        historical_price = get_historical_btc_price(payout.payout_date)

        # Update the payout's closing_price and fetched_at fields
        payout.closing_price = Decimal(str(historical_price))
        payout.closing_price_fetched_at = date.today()
        payout.save()

        return JsonResponse({
            'success': True,
            'closing_price': float(payout.closing_price),
            'formatted_price': f'${payout.closing_price:,.2f}'
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'Failed to fetch closing price: {str(e)}'
        })


# Background task state via Django cache (shareable across processes)
# For multi-process production, configure a shared cache backend (Redis, Memcached, or database)
# in settings.py. The default LocMemCache works for single-process dev servers.
from django.core.cache import cache

_BULK_FETCH_CACHE_KEY = 'bulk_fetch_closing_prices_status'
_API_FETCH_CACHE_KEY = 'api_fetch_status'
_CACHE_TIMEOUT = 3600  # 1 hour

_bulk_fetch_lock = threading.Lock()
_api_fetch_lock = threading.Lock()


def _get_bulk_fetch_status():
    return cache.get(_BULK_FETCH_CACHE_KEY, {
        'running': False, 'total': 0, 'processed': 0,
        'updated': 0, 'skipped': 0, 'errors': 0,
        'error_details': [], 'message': '',
    })


def _set_bulk_fetch_status(status):
    cache.set(_BULK_FETCH_CACHE_KEY, status, _CACHE_TIMEOUT)


def _get_api_fetch_status():
    return cache.get(_API_FETCH_CACHE_KEY, {
        'running': False, 'message': '', 'success': None,
    })


def _set_api_fetch_status(status):
    cache.set(_API_FETCH_CACHE_KEY, status, _CACHE_TIMEOUT)


def _bulk_fetch_closing_prices_task():
    """Background task: fetch closing prices in sub-batches with delay to respect API rate limits."""
    from .api_utils import get_historical_btc_price as fetch_price

    BATCH_SIZE = 5
    DELAY_BETWEEN_BATCHES = 3  # seconds

    today = date.today()

    payouts = list(
        Payout.objects.filter(payout_date__isnull=False).order_by('payout_date')
    )

    payouts_to_fetch = []
    for p in payouts:
        if p.closing_price_fetched_at is None:
            payouts_to_fetch.append(p)
        elif p.closing_price_fetched_at <= p.payout_date:
            payouts_to_fetch.append(p)

    status = _get_bulk_fetch_status()
    status.update({
        'total': len(payouts_to_fetch), 'processed': 0,
        'updated': 0, 'skipped': 0, 'errors': 0,
        'error_details': [],
        'message': f'Processing {len(payouts_to_fetch)} payouts...',
    })
    _set_bulk_fetch_status(status)

    for i in range(0, len(payouts_to_fetch), BATCH_SIZE):
        batch = payouts_to_fetch[i:i + BATCH_SIZE]

        for payout in batch:
            try:
                historical_price = fetch_price(payout.payout_date)
                payout.closing_price = Decimal(str(historical_price))
                payout.closing_price_fetched_at = today
                payout.save()
                status = _get_bulk_fetch_status()
                status['updated'] += 1
            except Exception as e:
                status = _get_bulk_fetch_status()
                status['errors'] += 1
                status['error_details'].append(
                    f'Payout #{payout.pk} ({payout.payout_date}): {str(e)}'
                )

            status['processed'] += 1
            _set_bulk_fetch_status(status)

        if i + BATCH_SIZE < len(payouts_to_fetch):
            time.sleep(DELAY_BETWEEN_BATCHES)

    status = _get_bulk_fetch_status()
    skipped = len(payouts) - len(payouts_to_fetch)
    status['skipped'] = skipped
    status['message'] = (
        f'Completed: {status["updated"]} updated, '
        f'{skipped} skipped, '
        f'{status["errors"]} errors.'
    )
    status['running'] = False
    _set_bulk_fetch_status(status)


@require_POST
def bulk_fetch_closing_prices(request):
    """Trigger bulk closing price fetch as a background task."""
    with _bulk_fetch_lock:
        status = _get_bulk_fetch_status()
        if status['running']:
            return JsonResponse({
                'success': False,
                'error': 'A bulk fetch is already in progress.'
            })
        status = {
            'running': True, 'total': 0, 'processed': 0,
            'updated': 0, 'skipped': 0, 'errors': 0,
            'error_details': [], 'message': 'Starting...',
        }
        _set_bulk_fetch_status(status)

    thread = threading.Thread(target=_bulk_fetch_closing_prices_task, daemon=True)
    thread.start()

    return JsonResponse({'success': True, 'message': 'Bulk fetch started.'})


def bulk_fetch_closing_prices_status(request):
    """Return the current status of the bulk closing price fetch task."""
    status = _get_bulk_fetch_status()
    return JsonResponse({
        'running': status['running'],
        'total': status['total'],
        'processed': status['processed'],
        'updated': status['updated'],
        'skipped': status['skipped'],
        'errors': status['errors'],
        'message': status['message'],
        'error_details': list(status['error_details']),
    })


def _fetch_api_data_task():
    """Background task: fetch all API data and save to database."""
    try:
        status = _get_api_fetch_status()
        status['message'] = 'Fetching API data...'
        _set_api_fetch_status(status)

        result = fetch_all_api_data()

        if result['success']:
            api_data = APIData.get_api_data()
            api_data.bitcoin_price_usd = result['bitcoin_price_usd']
            api_data.network_hashrate_ehs = result['network_hashrate_ehs']
            api_data.network_difficulty = result['network_difficulty']
            api_data.avg_block_fees_24h = result['avg_block_fees_24h']
            api_data.save()

            status = _get_api_fetch_status()
            status['message'] = result['message']
            status['success'] = True
        else:
            status = _get_api_fetch_status()
            status['message'] = result['message']
            status['success'] = False
    except Exception as e:
        status = _get_api_fetch_status()
        status['message'] = f'Unexpected error: {str(e)}'
        status['success'] = False
    finally:
        status['running'] = False
        _set_api_fetch_status(status)


@require_POST
def trigger_fetch_api_data(request):
    """Trigger API data fetch as a background task."""
    with _api_fetch_lock:
        status = _get_api_fetch_status()
        if status['running']:
            return JsonResponse({
                'success': False,
                'error': 'API fetch is already in progress.'
            })
        _set_api_fetch_status({
            'running': True, 'message': 'Starting...', 'success': None,
        })

    thread = threading.Thread(target=_fetch_api_data_task, daemon=True)
    thread.start()

    return JsonResponse({'success': True, 'message': 'API fetch started.'})


def fetch_api_data_status(request):
    """Return the current status of the API data fetch task."""
    status = _get_api_fetch_status()
    return JsonResponse({
        'running': status['running'],
        'message': status['message'],
        'success': status['success'],
    })


def api_data_view(request):
    """API Data page view"""
    api_data = APIData.get_api_data()
    return render(request, 'mining/api_data.html', {'api_data': api_data})


def settings_view(request):
    """Settings page view"""
    settings = Settings.get_settings()
    
    if request.method == 'POST':
        form = SettingsForm(request.POST, instance=settings)
        if form.is_valid():
            form.save()
            messages.success(request, 'Settings saved successfully!')
            return redirect('settings')
    else:
        form = SettingsForm(instance=settings)
    
    return render(request, 'mining/settings.html', {'form': form, 'settings': settings})


# Expense Views
class ExpenseListView(ListView):
    model = Expense
    template_name = 'mining/expense_list.html'
    context_object_name = 'expenses'
    paginate_by = 50
    queryset = Expense.objects.select_related('platform')


class ExpenseDetailView(DetailView):
    model = Expense
    template_name = 'mining/expense_detail.html'
    context_object_name = 'expense'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_expense = self.get_object()
        
        # Navigate by expense_date (matching list view order: newest first)
        # Previous = newer expense (appears before in list)
        previous_expense = Expense.objects.filter(
            Q(expense_date__gt=current_expense.expense_date) |
            Q(expense_date=current_expense.expense_date, id__gt=current_expense.id)
        ).order_by('expense_date', 'id').first()
        
        # Next = older expense (appears after in list)
        next_expense = Expense.objects.filter(
            Q(expense_date__lt=current_expense.expense_date) |
            Q(expense_date=current_expense.expense_date, id__lt=current_expense.id)
        ).order_by('-expense_date', '-id').first()
        
        context['previous_expense'] = previous_expense
        context['next_expense'] = next_expense
        return context


class ExpenseCreateView(CreateView):
    model = Expense
    form_class = ExpenseForm
    template_name = 'mining/expense_form.html'
    
    def get_success_url(self):
        return reverse_lazy('expense_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Expense created successfully!')
        return super().form_valid(form)


class ExpenseUpdateView(UpdateView):
    model = Expense
    form_class = ExpenseForm
    template_name = 'mining/expense_form.html'
    
    def get_success_url(self):
        return reverse_lazy('expense_detail', kwargs={'pk': self.object.pk})
    
    def form_valid(self, form):
        messages.success(self.request, 'Expense updated successfully!')
        return super().form_valid(form)


class ExpenseDeleteView(DeleteView):
    model = Expense
    template_name = 'mining/expense_confirm_delete.html'
    success_url = reverse_lazy('expense_list')
    context_object_name = 'expense'
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Expense deleted successfully!')
        return super().delete(request, *args, **kwargs)


# Dashboard Views
def overview_dashboard(request):
    """Overview Dashboard with comprehensive mining analytics"""
    selected_platform = resolve_selected_platform(request.GET.get('platform', ''))
    data = get_overview_data(selected_platform)
    return render(request, 'mining/overview_dashboard.html', data)


# Import Template Download Views
def download_platform_template(request):
    """Download import template for Remote Mining Platforms"""
    wb = Workbook()
    ws = wb.create_sheet(title='Platform Import Template')
    
    # Add headers based on form fields
    headers = ['name', 'website_link', 'portal_url', 'point_of_contact_name', 
               'point_of_contact_email', 'point_of_contact_phone', 'point_of_contact_telegram', 'energy_price']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="platform_import_template.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response


def download_miner_template(request):
    """Download import template for Miners"""
    wb = Workbook()
    ws = wb.create_sheet(title='Miner Import Template')
    
    # Add headers based on form fields (excluding image field for import)
    headers = ['model', 'manufacturer', 'product_link', 'serial_number', 
               'platform', 'platform_internal_id', 'hashrate', 'power', 'efficiency', 
               'purchase_price', 'purchase_date', 'start_date', 'location']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="miner_import_template.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response


def download_payout_template(request):
    """Download import template for Payouts"""
    wb = Workbook()
    ws = wb.create_sheet(title='Payout Import Template')
    
    # Add headers based on form fields
    headers = ['payout_date', 'payout_amount', 'platform', 'transaction_id', 'closing_price', 'value_at_payout (read-only)']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="payout_import_template.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response


# Data Export Views
def export_platform_data(request):
    """Export all platform data to Excel file"""
    wb = Workbook()
    ws = wb.create_sheet(title='Platform Data')
    
    # Add headers
    headers = ['name', 'website_link', 'portal_url', 'point_of_contact_name', 
               'point_of_contact_email', 'point_of_contact_phone', 
               'point_of_contact_telegram', 'energy_price']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    # Add data rows
    platforms = RemoteMiningPlatform.objects.all()
    for row, platform in enumerate(platforms, start=1):
        ws.cell(row=row + 1, column=1, value=platform.name or '')
        ws.cell(row=row + 1, column=2, value=platform.website_link or '')
        ws.cell(row=row + 1, column=3, value=platform.portal_url or '')
        ws.cell(row=row + 1, column=4, value=platform.point_of_contact_name or '')
        ws.cell(row=row + 1, column=5, value=platform.point_of_contact_email or '')
        ws.cell(row=row + 1, column=6, value=platform.point_of_contact_phone or '')
        ws.cell(row=row + 1, column=7, value=platform.point_of_contact_telegram or '')
        ws.cell(row=row + 1, column=8, value=float(platform.energy_price) if platform.energy_price else '')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="platform_data_export.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response


def export_miner_data(request):
    """Export all miner data to Excel file"""
    wb = Workbook()
    ws = wb.create_sheet(title='Miner Data')
    
    # Add headers
    headers = ['model', 'manufacturer', 'product_link', 'serial_number', 
               'platform', 'platform_internal_id', 'hashrate', 'power', 'efficiency', 
               'purchase_price', 'purchase_date', 'start_date', 'location']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    # Add data rows
    miners = Miner.objects.select_related('platform').all()
    for row, miner in enumerate(miners, start=1):
        ws.cell(row=row + 1, column=1, value=miner.model or '')
        ws.cell(row=row + 1, column=2, value=miner.manufacturer or '')
        ws.cell(row=row + 1, column=3, value=miner.product_link or '')
        ws.cell(row=row + 1, column=4, value=miner.serial_number or '')
        ws.cell(row=row + 1, column=5, value=miner.platform.pk if miner.platform else '')
        ws.cell(row=row + 1, column=6, value=miner.platform_internal_id or '')
        ws.cell(row=row + 1, column=7, value=float(miner.hashrate) if miner.hashrate else '')
        ws.cell(row=row + 1, column=8, value=float(miner.power) if miner.power else '')
        ws.cell(row=row + 1, column=9, value=float(miner.efficiency) if miner.efficiency else '')
        ws.cell(row=row + 1, column=10, value=float(miner.purchase_price) if miner.purchase_price else '')
        ws.cell(row=row + 1, column=11, value=miner.purchase_date.strftime('%Y-%m-%d') if miner.purchase_date else '')
        ws.cell(row=row + 1, column=12, value=miner.start_date.strftime('%Y-%m-%d') if miner.start_date else '')
        ws.cell(row=row + 1, column=13, value=miner.location or '')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="miner_data_export.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response


def export_payout_data(request):
    """Export all payout data to Excel file"""
    wb = Workbook()
    ws = wb.create_sheet(title='Payout Data')
    
    # Add headers
    headers = ['payout_date', 'payout_amount', 'platform', 'transaction_id', 'closing_price', 'value_at_payout']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    # Add data rows
    payouts = Payout.objects.select_related('platform').all()
    for row, payout in enumerate(payouts, start=1):
        ws.cell(row=row + 1, column=1, value=payout.payout_date.strftime('%Y-%m-%d'))
        ws.cell(row=row + 1, column=2, value=float(payout.payout_amount))
        ws.cell(row=row + 1, column=3, value=payout.platform.pk if payout.platform else '')
        ws.cell(row=row + 1, column=4, value=payout.transaction_id or '')
        ws.cell(row=row + 1, column=5, value=float(payout.closing_price) if payout.closing_price else '')
        ws.cell(row=row + 1, column=6, value=float(payout.value_at_payout) if payout.value_at_payout else '')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="payout_data_export.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response


def export_overview_data(request):
    """Export overview dashboard data to Excel file"""
    wb = Workbook()
    selected_platform = resolve_selected_platform(request.GET.get('platform', ''))
    data = get_overview_data(selected_platform)
    selected_platform_name = selected_platform.name if selected_platform else 'All Platforms'
    bitcoin_price = data['bitcoin_price']
    network_hashrate = data['network_hashrate']
    network_difficulty = data['network_difficulty']
    avg_block_fees_24h = data['avg_block_fees_24h']
    miner_count = data['miner_count']
    total_hashrate = data['total_hashrate']
    total_power = data['total_power']
    total_capex = data['total_capex']
    avg_efficiency = data['avg_efficiency']
    hashrate_weighted_efficiency = data['hashrate_weighted_efficiency']
    avg_energy_cost = data['avg_energy_cost']
    hashrate_weighted_energy_cost = data['hashrate_weighted_energy_cost']
    total_btc_mined = data['total_btc_mined']
    current_gross_value = data['current_gross_value']
    gross_value_at_payout = data['gross_value_at_payout']
    appreciation = data['appreciation']
    total_opex = data['total_opex']
    current_net_value = data['current_net_value']
    total_payouts = data['total_payouts']

    # Sheet 1: Overview Summary
    ws_summary = wb.create_sheet(title='Overview Summary')
    
    # Headers
    ws_summary.cell(row=1, column=1, value='Metric Category')
    ws_summary.cell(row=1, column=2, value='Metric Name')
    ws_summary.cell(row=1, column=3, value='Value')
    ws_summary.cell(row=1, column=4, value='Unit')
    
    # Data rows
    row = 1
    
    # Platform Filter
    ws_summary.cell(row=row + 1, column=1, value='Filter')
    ws_summary.cell(row=row + 1, column=2, value='Platform')
    ws_summary.cell(row=row + 1, column=3, value=selected_platform_name)
    ws_summary.cell(row=row + 1, column=4, value='')
    row += 1
    
    # Network Data
    ws_summary.cell(row=row + 1, column=1, value='Network Data')
    ws_summary.cell(row=row + 1, column=2, value='Bitcoin Spot Price')
    ws_summary.cell(row=row + 1, column=3, value=float(bitcoin_price))
    ws_summary.cell(row=row + 1, column=4, value='USD')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Network Data')
    ws_summary.cell(row=row + 1, column=2, value='Total Network Hashrate')
    ws_summary.cell(row=row + 1, column=3, value=float(network_hashrate))
    ws_summary.cell(row=row + 1, column=4, value='EH/s')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Network Data')
    ws_summary.cell(row=row + 1, column=2, value='Network Difficulty')
    ws_summary.cell(row=row + 1, column=3, value=float(network_difficulty))
    ws_summary.cell(row=row + 1, column=4, value='')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Network Data')
    ws_summary.cell(row=row + 1, column=2, value='24h Avg Block Fees')
    ws_summary.cell(row=row + 1, column=3, value=float(avg_block_fees_24h))
    ws_summary.cell(row=row + 1, column=4, value='BTC')
    row += 1
    
    # Fleet Data
    ws_summary.cell(row=row + 1, column=1, value='Fleet Data')
    ws_summary.cell(row=row + 1, column=2, value='Miner Count')
    ws_summary.cell(row=row + 1, column=3, value=miner_count)
    ws_summary.cell(row=row + 1, column=4, value='units')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Fleet Data')
    ws_summary.cell(row=row + 1, column=2, value='Total Hashrate')
    ws_summary.cell(row=row + 1, column=3, value=float(total_hashrate))
    ws_summary.cell(row=row + 1, column=4, value='TH/s')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Fleet Data')
    ws_summary.cell(row=row + 1, column=2, value='Total Power')
    ws_summary.write(row, 2, round(float(total_power), 2))  # Power already stored in kW in database
    ws_summary.cell(row=row + 1, column=4, value='kW')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Fleet Data')
    ws_summary.cell(row=row + 1, column=2, value='Total Hardware Cost')
    ws_summary.cell(row=row + 1, column=3, value=float(total_capex))
    ws_summary.cell(row=row + 1, column=4, value='USD')
    row += 1
    
    # Efficiency Data
    ws_summary.cell(row=row + 1, column=1, value='Efficiency Data')
    ws_summary.cell(row=row + 1, column=2, value='Average Efficiency')
    ws_summary.cell(row=row + 1, column=3, value=float(avg_efficiency))
    ws_summary.cell(row=row + 1, column=4, value='W/TH')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Efficiency Data')
    ws_summary.cell(row=row + 1, column=2, value='Hashrate Weighted Avg Efficiency')
    ws_summary.cell(row=row + 1, column=3, value=float(hashrate_weighted_efficiency))
    ws_summary.cell(row=row + 1, column=4, value='W/TH')
    row += 1
    
    # Energy Data
    ws_summary.cell(row=row + 1, column=1, value='Energy Data')
    ws_summary.cell(row=row + 1, column=2, value='Average Energy Cost')
    ws_summary.cell(row=row + 1, column=3, value=float(avg_energy_cost))
    ws_summary.cell(row=row + 1, column=4, value='$/kWh')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Energy Data')
    ws_summary.cell(row=row + 1, column=2, value='Hashrate Weighted Avg Energy Cost')
    ws_summary.cell(row=row + 1, column=3, value=float(hashrate_weighted_energy_cost))
    ws_summary.cell(row=row + 1, column=4, value='$/kWh')
    row += 1
    
    # Revenue Data
    ws_summary.cell(row=row + 1, column=1, value='Revenue Data')
    ws_summary.cell(row=row + 1, column=2, value='Total BTC Mined')
    ws_summary.cell(row=row + 1, column=3, value=float(total_btc_mined))
    ws_summary.cell(row=row + 1, column=4, value='BTC')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Revenue Data')
    ws_summary.cell(row=row + 1, column=2, value='Current Gross Value')
    ws_summary.cell(row=row + 1, column=3, value=round(float(current_gross_value), 2))
    ws_summary.cell(row=row + 1, column=4, value='USD')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Revenue Data')
    ws_summary.cell(row=row + 1, column=2, value='Gross Value at Payout')
    ws_summary.cell(row=row + 1, column=3, value=round(float(gross_value_at_payout), 2))
    ws_summary.cell(row=row + 1, column=4, value='USD')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Revenue Data')
    ws_summary.cell(row=row + 1, column=2, value='Appreciation')
    ws_summary.cell(row=row + 1, column=3, value=round(float(appreciation), 2))
    ws_summary.cell(row=row + 1, column=4, value='USD')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Revenue Data')
    ws_summary.cell(row=row + 1, column=2, value='Total Payouts')
    ws_summary.cell(row=row + 1, column=3, value=total_payouts)
    ws_summary.cell(row=row + 1, column=4, value='count')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Revenue Data')
    ws_summary.cell(row=row + 1, column=2, value='Total OPEX')
    ws_summary.cell(row=row + 1, column=3, value=round(float(total_opex), 2))
    ws_summary.cell(row=row + 1, column=4, value='USD')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Revenue Data')
    ws_summary.cell(row=row + 1, column=2, value='Current Net Value')
    ws_summary.cell(row=row + 1, column=3, value=round(float(current_net_value), 2))
    ws_summary.cell(row=row + 1, column=4, value='USD')
    row += 1
    
    # Sheet 2: Hashrate by Platform
    ws_hashrate_platform = wb.create_sheet(title='Hashrate by Platform')
    ws_hashrate_platform.cell(row=1, column=1, value='Platform')
    ws_hashrate_platform.cell(row=1, column=2, value='Hashrate (TH/s)')

    platform_row = 1
    for item in data['hashrate_by_platform']:
        ws_hashrate_platform.cell(row=platform_row + 1, column=1, value=item['platform'])
        ws_hashrate_platform.cell(row=platform_row + 1, column=2, value=item['hashrate'])
        platform_row += 1

    # Sheet 3: Hashrate by Location
    ws_hashrate_location = wb.create_sheet(title='Hashrate by Location')
    ws_hashrate_location.cell(row=1, column=1, value='Location')
    ws_hashrate_location.cell(row=1, column=2, value='Hashrate (TH/s)')

    location_row = 1
    for item in data['hashrate_by_location']:
        ws_hashrate_location.cell(row=location_row + 1, column=1, value=item['location'])
        ws_hashrate_location.cell(row=location_row + 1, column=2, value=item['hashrate'])
        location_row += 1

    # Sheet 4: Revenue by Platform
    ws_revenue_platform = wb.create_sheet(title='Revenue by Platform')
    ws_revenue_platform.cell(row=1, column=1, value='Platform')
    ws_revenue_platform.cell(row=1, column=2, value='BTC Mined')
    ws_revenue_platform.cell(row=1, column=3, value='Gross Value (USD)')
    ws_revenue_platform.cell(row=1, column=4, value='Gross Value at Payout (USD)')
    ws_revenue_platform.cell(row=1, column=5, value='Payout Count')

    revenue_row = 1
    for item in data['revenue_by_platform']:
        ws_revenue_platform.cell(row=revenue_row + 1, column=1, value=item['platform'])
        ws_revenue_platform.cell(row=revenue_row + 1, column=2, value=item['btc_mined'])
        ws_revenue_platform.cell(row=revenue_row + 1, column=3, value=round(item['gross_value'], 2))
        ws_revenue_platform.cell(row=revenue_row + 1, column=4, value=round(item['gross_value_at_payout'], 2))
        ws_revenue_platform.cell(row=revenue_row + 1, column=5, value=item['payout_count'])
        revenue_row += 1
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    platform_suffix = f'_{selected_platform_name.replace(" ", "_")}' if selected_platform else ''
    response['Content-Disposition'] = f'attachment; filename="overview_dashboard{platform_suffix}_export.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response


# Data Import Views
def import_platform_data(request):
    """Import platform data from uploaded Excel file"""
    if request.method == 'POST' and request.FILES.get('import_file'):
        try:
            file = request.FILES['import_file']
            wb = load_workbook(file)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(str(header).strip())

            # Process data rows
            imported_count = 0
            valid_fields = {f.name for f in RemoteMiningPlatform._meta.get_fields()}
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers):
                    cell_value = ws.cell(row=row, column=col_idx + 1).value
                    if cell_value is not None:
                        if isinstance(cell_value, datetime):
                            row_data[header] = cell_value.date()
                        elif isinstance(cell_value, str):
                            row_data[header] = cell_value.strip()
                        else:
                            row_data[header] = cell_value

                if row_data:
                    platform_data = {}
                    for field, value in row_data.items():
                        if field in valid_fields and value:
                            if field == 'energy_price':
                                platform_data[field] = Decimal(str(value))
                            else:
                                platform_data[field] = value

                    if platform_data:
                        RemoteMiningPlatform.objects.create(**platform_data)
                        imported_count += 1

            messages.success(request, f'Successfully imported {imported_count} platforms!')
            return redirect('platform_list')
            
        except Exception as e:
            messages.error(request, 'Wrong import file format or data. Please check your file and try again.')
            return redirect('platform_list')
    
    return redirect('platform_list')


def import_miner_data(request):
    """Import miner data from uploaded Excel file"""
    if request.method == 'POST' and request.FILES.get('import_file'):
        try:
            file = request.FILES['import_file']
            wb = load_workbook(file)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(str(header).strip())

            # Process data rows
            imported_count = 0
            valid_fields = {f.name for f in Miner._meta.get_fields()}
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers):
                    cell_value = ws.cell(row=row, column=col_idx + 1).value
                    if cell_value is not None:
                        if isinstance(cell_value, datetime):
                            row_data[header] = cell_value.date()
                        elif isinstance(cell_value, str):
                            row_data[header] = cell_value.strip()
                        else:
                            row_data[header] = cell_value

                if row_data:
                    miner_data = {}
                    for field, value in row_data.items():
                        if field in valid_fields and value:
                            if field == 'platform':
                                try:
                                    platform = RemoteMiningPlatform.objects.get(pk=int(float(value)))
                                    miner_data[field] = platform
                                except (RemoteMiningPlatform.DoesNotExist, ValueError, TypeError):
                                    continue
                            elif field in ['hashrate', 'power', 'efficiency', 'purchase_price']:
                                miner_data[field] = Decimal(str(value))
                            elif field in ['purchase_date', 'start_date']:
                                if isinstance(value, str):
                                    miner_data[field] = datetime.strptime(value, '%Y-%m-%d').date()
                                else:
                                    miner_data[field] = value
                            else:
                                miner_data[field] = value

                    if miner_data:
                        Miner.objects.create(**miner_data)
                        imported_count += 1

            messages.success(request, f'Successfully imported {imported_count} miners!')
            return redirect('miner_list')
            
        except Exception as e:
            messages.error(request, 'Wrong import file format or data. Please check your file and try again.')
            return redirect('miner_list')
    
    return redirect('miner_list')


def import_payout_data(request):
    """Import payout data from uploaded Excel file"""
    if request.method == 'POST' and request.FILES.get('import_file'):
        try:
            file = request.FILES['import_file']
            wb = load_workbook(file)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(str(header).strip())

            # Process data rows
            imported_count = 0
            valid_fields = {f.name for f in Payout._meta.get_fields()}
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers):
                    cell_value = ws.cell(row=row, column=col_idx + 1).value
                    if cell_value is not None:
                        if isinstance(cell_value, datetime):
                            row_data[header] = cell_value.date()
                        elif isinstance(cell_value, str):
                            row_data[header] = cell_value.strip()
                        else:
                            row_data[header] = cell_value

                if row_data:
                    payout_data = {}
                    for field, value in row_data.items():
                        if field in valid_fields and value:
                            if field == 'platform':
                                try:
                                    platform = RemoteMiningPlatform.objects.get(pk=int(float(value)))
                                    payout_data[field] = platform
                                except (RemoteMiningPlatform.DoesNotExist, ValueError, TypeError):
                                    continue
                            elif field in ['payout_amount', 'closing_price']:
                                payout_data[field] = Decimal(str(value))
                            elif field == 'payout_date':
                                if isinstance(value, str):
                                    payout_data[field] = datetime.strptime(value, '%Y-%m-%d').date()
                                else:
                                    payout_data[field] = value
                            else:
                                payout_data[field] = value

                    if payout_data:
                        Payout.objects.create(**payout_data)
                        imported_count += 1

            messages.success(request, f'Successfully imported {imported_count} payouts!')
            return redirect('payout_list')
            
        except Exception as e:
            messages.error(request, 'Wrong import file format or data. Please check your file and try again.')
            return redirect('payout_list')
    
    return redirect('payout_list')


# Expense Import/Export Functions
def download_expense_template(request):
    """Download import template for Expenses"""
    wb = Workbook()
    ws = wb.create_sheet(title='Expense Import Template')
    
    # Add headers based on form fields
    headers = ['expense_date', 'platform', 'category', 'description', 'expense_amount', 'invoice_link', 'receipt_link', 'notes']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="expense_import_template.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response


def export_expense_data(request):
    """Export all expense data to Excel file"""
    wb = Workbook()
    ws = wb.create_sheet(title='Expense Data')
    
    # Add headers
    headers = ['expense_date', 'platform', 'category', 'description', 'expense_amount', 'invoice_link', 'receipt_link', 'notes']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    # Add data rows
    expenses = Expense.objects.select_related('platform').all().order_by('-expense_date')
    for row, expense in enumerate(expenses, start=1):
        ws.cell(row=row + 1, column=1, value=expense.expense_date.strftime('%Y-%m-%d') if expense.expense_date else '')
        ws.cell(row=row + 1, column=2, value=expense.platform.id if expense.platform else '')
        ws.cell(row=row + 1, column=3, value=expense.category or '')
        ws.cell(row=row + 1, column=4, value=expense.description or '')
        ws.cell(row=row + 1, column=5, value=float(expense.expense_amount) if expense.expense_amount else '')
        ws.cell(row=row + 1, column=6, value=expense.invoice_link or '')
        ws.cell(row=row + 1, column=7, value=expense.receipt_link or '')
        ws.cell(row=row + 1, column=8, value=expense.notes or '')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="expense_data_export.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response


def import_expense_data(request):
    """Import expense data from uploaded Excel file"""
    if request.method == 'POST' and request.FILES.get('import_file'):
        try:
            file = request.FILES['import_file']
            wb = load_workbook(file)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(str(header).lower().strip())

            # Process data rows
            imported_count = 0
            for row in range(2, ws.max_row + 1):
                expense_data = {}

                for col_idx, header in enumerate(headers):
                    cell_value = ws.cell(row=row, column=col_idx + 1).value

                    if header == 'expense_date' and cell_value:
                        try:
                            if isinstance(cell_value, datetime):
                                expense_data['expense_date'] = cell_value.date()
                            else:
                                expense_data['expense_date'] = datetime.strptime(str(cell_value), '%Y-%m-%d').date()
                        except (ValueError, TypeError) as e:
                            logger.warning("Expense import: bad date at row %d: %s", row, e)
                            continue
                    elif header == 'platform' and cell_value:
                        try:
                            platform_id = int(float(cell_value))
                            platform = RemoteMiningPlatform.objects.get(pk=platform_id)
                            expense_data['platform'] = platform
                        except (RemoteMiningPlatform.DoesNotExist, ValueError, TypeError):
                            pass
                    elif header == 'category' and cell_value:
                        category_value = str(cell_value).upper().strip()
                        if category_value in ['CAPEX', 'OPEX']:
                            expense_data['category'] = category_value
                    elif header == 'description' and cell_value:
                        expense_data['description'] = str(cell_value)
                    elif header == 'expense_amount' and cell_value:
                        try:
                            expense_data['expense_amount'] = Decimal(str(cell_value))
                        except (ValueError, InvalidOperation):
                            pass
                    elif header == 'invoice_link' and cell_value:
                        expense_data['invoice_link'] = str(cell_value)
                    elif header == 'receipt_link' and cell_value:
                        expense_data['receipt_link'] = str(cell_value)
                    elif header == 'notes' and cell_value:
                        expense_data['notes'] = str(cell_value)
                
                # Create expense if we have required fields
                if 'expense_date' in expense_data and 'category' in expense_data and 'expense_amount' in expense_data:
                    Expense.objects.create(**expense_data)
                    imported_count += 1
            
            messages.success(request, f'Successfully imported {imported_count} expenses!')
            return redirect('expense_list')
            
        except Exception as e:
            messages.error(request, 'Wrong import file format or data. Please check your file and try again.')
            return redirect('expense_list')
    
    return redirect('expense_list')


# ===== TOP-UP VIEWS =====

class TopUpListView(ListView):
    model = TopUp
    template_name = 'mining/topup_list.html'
    context_object_name = 'topups'
    paginate_by = 50
    queryset = TopUp.objects.select_related('platform')


class TopUpDetailView(DetailView):
    model = TopUp
    template_name = 'mining/topup_detail.html'
    context_object_name = 'topup'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Navigate by topup_date (matching list view order: newest first)
        topup = self.get_object()
        # Previous = newer top-up (appears before in list)
        context['previous_topup'] = TopUp.objects.filter(
            Q(topup_date__gt=topup.topup_date) |
            Q(topup_date=topup.topup_date, id__gt=topup.id)
        ).order_by('topup_date', 'id').first()
        
        # Next = older top-up (appears after in list)
        context['next_topup'] = TopUp.objects.filter(
            Q(topup_date__lt=topup.topup_date) |
            Q(topup_date=topup.topup_date, id__lt=topup.id)
        ).order_by('-topup_date', '-id').first()
        
        return context


class TopUpCreateView(CreateView):
    model = TopUp
    form_class = TopUpForm
    template_name = 'mining/topup_form.html'
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Top-Up created successfully!')
        return response
    
    def get_success_url(self):
        return reverse('topup_detail', kwargs={'pk': self.object.pk})


class TopUpUpdateView(UpdateView):
    model = TopUp
    form_class = TopUpForm
    template_name = 'mining/topup_form.html'
    
    def form_valid(self, form):
        response = super().form_valid(form)
        messages.success(self.request, 'Top-Up updated successfully!')
        return response
    
    def get_success_url(self):
        return reverse('topup_detail', kwargs={'pk': self.object.pk})


class TopUpDeleteView(DeleteView):
    model = TopUp
    template_name = 'mining/topup_confirm_delete.html'
    
    def delete(self, request, *args, **kwargs):
        response = super().delete(request, *args, **kwargs)
        messages.success(request, 'Top-Up deleted successfully!')
        return response
    
    def get_success_url(self):
        return reverse('topup_list')


def download_topup_template(request):
    """Download Excel template for Top-Up import"""
    wb = Workbook()
    ws = wb.create_sheet(title='TopUp Template')
    
    # Add headers
    headers = ['topup_date', 'platform', 'topup_amount', 'description', 'receipt_link']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="topup_import_template.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response


def export_topup_data(request):
    """Export all top-up data to Excel file"""
    wb = Workbook()
    ws = wb.create_sheet(title='TopUp Data')
    
    # Add headers
    headers = ['topup_date', 'platform', 'topup_amount', 'description', 'receipt_link']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    # Add data rows
    topups = TopUp.objects.select_related('platform').all().order_by('-topup_date')
    for row, topup in enumerate(topups, start=1):
        ws.cell(row=row + 1, column=1, value=str(topup.topup_date) if topup.topup_date else '')
        ws.cell(row=row + 1, column=2, value=topup.platform.name if topup.platform else '')
        ws.cell(row=row + 1, column=3, value=float(topup.topup_amount) if topup.topup_amount else '')
        ws.cell(row=row + 1, column=4, value=topup.description or '')
        ws.cell(row=row + 1, column=5, value=topup.receipt_link or '')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="topup_data_export.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response


def import_topup_data(request):
    """Import top-up data from uploaded Excel file"""
    if request.method == 'POST' and request.FILES.get('import_file'):
        try:
            file = request.FILES['import_file']
            wb = load_workbook(file)
            ws = wb.active
            
            # Get headers from first row
            headers = [str(ws.cell(row=1, column=col).value).lower().strip() for col in range(1, ws.max_column + 1)]

            imported_count = 0

            # Process each row (skip header row)
            for row in range(2, ws.max_row + 1):
                topup_data = {}

                for col_idx, header in enumerate(headers):
                    cell_value = ws.cell(row=row, column=col_idx + 1).value

                    if header == 'topup_date' and cell_value:
                        try:
                            if isinstance(cell_value, datetime):
                                topup_data['topup_date'] = cell_value.date()
                            else:
                                topup_data['topup_date'] = datetime.strptime(str(cell_value), '%Y-%m-%d').date()
                        except (ValueError, TypeError) as e:
                            logger.warning("Top-up import: bad date at row %d: %s", row, e)
                            continue
                    elif header == 'platform' and cell_value:
                        try:
                            # Try to find platform by ID or name
                            if isinstance(cell_value, float):
                                platform = RemoteMiningPlatform.objects.get(id=int(cell_value))
                            else:
                                platform = RemoteMiningPlatform.objects.get(name=str(cell_value))
                            topup_data['platform'] = platform
                        except RemoteMiningPlatform.DoesNotExist:
                            continue
                    elif header == 'topup_amount' and cell_value:
                        try:
                            topup_data['topup_amount'] = float(cell_value)
                        except (ValueError, TypeError):
                            continue
                    elif header == 'description' and cell_value:
                        topup_data['description'] = str(cell_value)
                    elif header == 'receipt_link' and cell_value:
                        topup_data['receipt_link'] = str(cell_value)
                
                # Create top-up if we have required fields
                if 'topup_date' in topup_data and 'platform' in topup_data and 'topup_amount' in topup_data:
                    TopUp.objects.create(**topup_data)
                    imported_count += 1
            
            messages.success(request, f'Successfully imported {imported_count} top-ups!')
            return redirect('topup_list')
            
        except Exception as e:
            messages.error(request, 'Wrong import file format or data. Please check your file and try again.')
            return redirect('topup_list')
    
    return redirect('topup_list')


def forecasting_dashboard(request):
    """Forecasting Dashboard with BTC mining profitability calculations"""
    selected_platform = resolve_selected_platform(request.GET.get('platform', ''))
    data = get_forecasting_data(selected_platform)
    return render(request, 'mining/forecasting_dashboard.html', data)


def export_forecasting_data(request):
    """Export forecasting dashboard data to Excel file"""
    wb = Workbook()
    selected_platform = resolve_selected_platform(request.GET.get('platform', ''))
    data = get_forecasting_data(selected_platform)
    selected_platform_name = selected_platform.name if selected_platform else 'All Platforms'
    total_hashrate = data['total_hashrate']
    miner_count = data['miner_count']
    total_miner_count = data['total_miner_count']
    network_difficulty = data['network_difficulty']
    network_hashrate_ehs = data['network_hashrate_ehs']
    avg_tx_fees = data['avg_tx_fees']
    pool_fee = data['pool_fee']
    btc_price_usd = data['btc_price_usd']
    avg_efficiency = data['avg_efficiency']
    hashrate_weighted_efficiency = data['hashrate_weighted_efficiency']
    avg_energy_cost = data['avg_energy_cost']
    hashrate_weighted_energy_cost = data['hashrate_weighted_energy_cost']
    results = data['results']

    # Check if we have required data
    if results is None:
        ws = wb.create_sheet(title='Error')
        ws.cell(row=1, column=1, value='Error: Insufficient data (missing miners, API data, or network difficulty)')
    else:
        power_watts = results['power_consumption_watts']
        margin = results['margin']
        energy_cost_percentage = results['energy_cost_percentage']
        hashrate_share_percent = results['hashrate_share_percent']
        days_to_mine_1_btc = results['days_to_mine_1_btc']
        days_to_mine_small_btc = results['days_to_mine_small_btc']
        roi_data = results['roi_data']

        def format_time_breakdown(days_total):
            if days_total == float('inf'):
                return 'Never (not profitable)'
            years = int(days_total / 365)
            months = int((days_total % 365) / 30)
            remaining_days = int(days_total % 30)
            return f"{years} years, {months} months, {remaining_days} days"

        time_to_mine_1_btc = format_time_breakdown(days_to_mine_1_btc)
        time_to_mine_small_btc = format_time_breakdown(days_to_mine_small_btc)

        # Sheet 1: Forecasting Summary
        ws_summary = wb.create_sheet(title='Forecasting Summary')
        ws_summary.cell(row=1, column=1, value='Metric Category')
        ws_summary.cell(row=1, column=2, value='Metric Name')
        ws_summary.cell(row=1, column=3, value='Value')
        ws_summary.cell(row=1, column=4, value='Unit')
        
        row = 1
        
        # Platform Filter
        ws_summary.cell(row=row + 1, column=1, value='Filter')
        ws_summary.cell(row=row + 1, column=2, value='Platform')
        ws_summary.cell(row=row + 1, column=3, value=selected_platform_name)
        ws_summary.cell(row=row + 1, column=4, value='')
        row += 1
        
        # Network Overview
        ws_summary.cell(row=row + 1, column=1, value='Network Overview')
        ws_summary.cell(row=row + 1, column=2, value='Network Difficulty')
        ws_summary.cell(row=row + 1, column=3, value=network_difficulty)
        ws_summary.cell(row=row + 1, column=4, value='')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='Network Overview')
        ws_summary.cell(row=row + 1, column=2, value='Network Hashrate')
        ws_summary.cell(row=row + 1, column=3, value=network_hashrate_ehs)
        ws_summary.cell(row=row + 1, column=4, value='EH/s')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='Network Overview')
        ws_summary.cell(row=row + 1, column=2, value='Avg Block Fees (24h)')
        ws_summary.cell(row=row + 1, column=3, value=round(avg_tx_fees, 8))
        ws_summary.cell(row=row + 1, column=4, value='BTC')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='Network Overview')
        ws_summary.cell(row=row + 1, column=2, value='BTC Price')
        ws_summary.cell(row=row + 1, column=3, value=round(btc_price_usd, 2))
        ws_summary.cell(row=row + 1, column=4, value='USD')
        row += 1
        
        # My Fleet Overview
        ws_summary.cell(row=row + 1, column=1, value='My Fleet Overview')
        ws_summary.cell(row=row + 1, column=2, value='Miners Accounted For')
        ws_summary.cell(row=row + 1, column=3, value=miner_count)
        ws_summary.cell(row=row + 1, column=4, value=f'of {total_miner_count}')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='My Fleet Overview')
        ws_summary.cell(row=row + 1, column=2, value='My Hashrate')
        ws_summary.cell(row=row + 1, column=3, value=float(total_hashrate))
        ws_summary.cell(row=row + 1, column=4, value='TH/s')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='My Fleet Overview')
        ws_summary.cell(row=row + 1, column=2, value='Pool Fee')
        ws_summary.cell(row=row + 1, column=3, value=pool_fee)
        ws_summary.cell(row=row + 1, column=4, value='%')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='My Fleet Overview')
        ws_summary.cell(row=row + 1, column=2, value='Power Consumption')
        ws_summary.cell(row=row + 1, column=3, value=round(power_watts / 1000, 2))
        ws_summary.cell(row=row + 1, column=4, value='kW')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='My Fleet Overview')
        ws_summary.cell(row=row + 1, column=2, value='Net Profit Margin')
        ws_summary.cell(row=row + 1, column=3, value=round(margin, 2))
        ws_summary.cell(row=row + 1, column=4, value='%')
        row += 1
        
        # Efficiency Data
        ws_summary.cell(row=row + 1, column=1, value='Efficiency Data')
        ws_summary.cell(row=row + 1, column=2, value='Average Efficiency')
        ws_summary.cell(row=row + 1, column=3, value=float(avg_efficiency))
        ws_summary.cell(row=row + 1, column=4, value='W/TH')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='Efficiency Data')
        ws_summary.cell(row=row + 1, column=2, value='Hashrate Weighted Avg Efficiency')
        ws_summary.cell(row=row + 1, column=3, value=round(float(hashrate_weighted_efficiency), 2))
        ws_summary.cell(row=row + 1, column=4, value='W/TH')
        row += 1
        
        # Energy Data
        ws_summary.cell(row=row + 1, column=1, value='Energy Data')
        ws_summary.cell(row=row + 1, column=2, value='Average Energy Cost')
        ws_summary.cell(row=row + 1, column=3, value=float(avg_energy_cost))
        ws_summary.cell(row=row + 1, column=4, value='$/kWh')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='Energy Data')
        ws_summary.cell(row=row + 1, column=2, value='Hashrate Weighted Avg Energy Cost')
        ws_summary.cell(row=row + 1, column=3, value=round(float(hashrate_weighted_energy_cost), 6))
        ws_summary.cell(row=row + 1, column=4, value='$/kWh')
        row += 1
        
        # Key Metrics
        ws_summary.cell(row=row + 1, column=1, value='Key Metrics')
        ws_summary.cell(row=row + 1, column=2, value='Time to mine 1 BTC')
        ws_summary.cell(row=row + 1, column=3, value=time_to_mine_1_btc)
        ws_summary.cell(row=row + 1, column=4, value='')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='Key Metrics')
        ws_summary.cell(row=row + 1, column=2, value='Time to mine 0.005 BTC')
        ws_summary.cell(row=row + 1, column=3, value=time_to_mine_small_btc)
        ws_summary.cell(row=row + 1, column=4, value='')
        row += 1
        
        if roi_data:
            roi_display = f"{roi_data['time_breakdown']['years']} years, {roi_data['time_breakdown']['months']} months, {roi_data['time_breakdown']['days']} days" if roi_data['days_to_roi'] != float('inf') else "Never (not profitable)"
            ws_summary.cell(row=row + 1, column=1, value='Key Metrics')
            ws_summary.cell(row=row + 1, column=2, value='ROI')
            ws_summary.cell(row=row + 1, column=3, value=roi_display)
            ws_summary.cell(row=row + 1, column=4, value='')
            row += 1
        
        # Sheet 2: Daily Projections
        ws_daily = wb.create_sheet(title='Daily Projections')
        ws_daily.cell(row=1, column=1, value='Projection Type')
        ws_daily.cell(row=1, column=2, value='BTC Amount')
        ws_daily.cell(row=1, column=3, value='USD Amount')
        
        row = 1
        ws_daily.cell(row=row + 1, column=1, value='Gross Payout (before pool fee)')
        ws_daily.cell(row=row + 1, column=2, value=round(daily_btc_gross_before_fee, 8))
        ws_daily.cell(row=row + 1, column=3, value=round(daily_usd_gross, 2))
        row += 1
        
        ws_daily.cell(row=row + 1, column=1, value='Pool Fee')
        ws_daily.cell(row=row + 1, column=2, value=round(pool_fee_btc, 8))
        ws_daily.cell(row=row + 1, column=3, value=round(pool_fee_btc * btc_price_usd, 2))
        row += 1
        
        ws_daily.cell(row=row + 1, column=1, value='After Pool Fee')
        ws_daily.cell(row=row + 1, column=2, value=round(daily_btc_after_fee, 8))
        ws_daily.cell(row=row + 1, column=3, value=round(daily_usd_after_fee, 2))
        row += 1
        
        ws_daily.cell(row=row + 1, column=1, value='Electricity Cost')
        ws_daily.cell(row=row + 1, column=2, value=round(daily_electricity_cost_usd / btc_price_usd, 8) if btc_price_usd > 0 else 0)
        ws_daily.cell(row=row + 1, column=3, value=round(daily_electricity_cost_usd, 2))
        row += 1
        
        ws_daily.cell(row=row + 1, column=1, value='Net Profit')
        ws_daily.cell(row=row + 1, column=2, value=round(daily_btc_net, 8))
        ws_daily.cell(row=row + 1, column=3, value=round(daily_usd_net, 2))
        row += 1
        
        # Sheet 3: Monthly Projections
        ws_monthly = wb.create_sheet(title='Monthly Projections')
        ws_monthly.cell(row=1, column=1, value='Projection Type')
        ws_monthly.cell(row=1, column=2, value='BTC Amount')
        ws_monthly.cell(row=1, column=3, value='USD Amount')
        
        row = 1
        ws_monthly.cell(row=row + 1, column=1, value='Gross Payout (before pool fee)')
        ws_monthly.cell(row=row + 1, column=2, value=round(daily_btc_gross_before_fee * 30, 8))
        ws_monthly.cell(row=row + 1, column=3, value=round(daily_usd_gross * 30, 2))
        row += 1
        
        ws_monthly.cell(row=row + 1, column=1, value='Pool Fee')
        ws_monthly.cell(row=row + 1, column=2, value=round(pool_fee_btc * 30, 8))
        ws_monthly.cell(row=row + 1, column=3, value=round(pool_fee_btc * btc_price_usd * 30, 2))
        row += 1
        
        ws_monthly.cell(row=row + 1, column=1, value='After Pool Fee')
        ws_monthly.cell(row=row + 1, column=2, value=round(daily_btc_after_fee * 30, 8))
        ws_monthly.cell(row=row + 1, column=3, value=round(daily_usd_after_fee * 30, 2))
        row += 1
        
        ws_monthly.cell(row=row + 1, column=1, value='Electricity Cost')
        ws_monthly.cell(row=row + 1, column=2, value=round(daily_electricity_cost_usd / btc_price_usd * 30, 8) if btc_price_usd > 0 else 0)
        ws_monthly.cell(row=row + 1, column=3, value=round(daily_electricity_cost_usd * 30, 2))
        row += 1
        
        ws_monthly.cell(row=row + 1, column=1, value='Net Profit')
        ws_monthly.cell(row=row + 1, column=2, value=round(daily_btc_net * 30, 8))
        ws_monthly.cell(row=row + 1, column=3, value=round(daily_usd_net * 30, 2))
        row += 1
        
        # Sheet 4: Yearly Projections
        ws_yearly = wb.create_sheet(title='Yearly Projections')
        ws_yearly.cell(row=1, column=1, value='Projection Type')
        ws_yearly.cell(row=1, column=2, value='BTC Amount')
        ws_yearly.cell(row=1, column=3, value='USD Amount')
        
        row = 1
        ws_yearly.cell(row=row + 1, column=1, value='Gross Payout (before pool fee)')
        ws_yearly.cell(row=row + 1, column=2, value=round(daily_btc_gross_before_fee * 365, 8))
        ws_yearly.cell(row=row + 1, column=3, value=round(daily_usd_gross * 365, 2))
        row += 1
        
        ws_yearly.cell(row=row + 1, column=1, value='Pool Fee')
        ws_yearly.cell(row=row + 1, column=2, value=round(pool_fee_btc * 365, 8))
        ws_yearly.cell(row=row + 1, column=3, value=round(pool_fee_btc * btc_price_usd * 365, 2))
        row += 1
        
        ws_yearly.cell(row=row + 1, column=1, value='After Pool Fee')
        ws_yearly.cell(row=row + 1, column=2, value=round(daily_btc_after_fee * 365, 8))
        ws_yearly.cell(row=row + 1, column=3, value=round(daily_usd_after_fee * 365, 2))
        row += 1
        
        ws_yearly.cell(row=row + 1, column=1, value='Electricity Cost')
        ws_yearly.cell(row=row + 1, column=2, value=round(daily_electricity_cost_usd / btc_price_usd * 365, 8) if btc_price_usd > 0 else 0)
        ws_yearly.cell(row=row + 1, column=3, value=round(daily_electricity_cost_usd * 365, 2))
        row += 1
        
        ws_yearly.cell(row=row + 1, column=1, value='Net Profit')
        ws_yearly.cell(row=row + 1, column=2, value=round(daily_btc_net * 365, 8))
        ws_yearly.cell(row=row + 1, column=3, value=round(daily_usd_net * 365, 2))
        row += 1
        
        # Sheet 5: Cost Basis Analysis
        ws_cost = wb.create_sheet(title='Cost Basis Analysis')
        ws_cost.cell(row=1, column=1, value='Cost Analysis')
        ws_cost.cell(row=1, column=2, value='Value')
        ws_cost.cell(row=1, column=3, value='Unit')
        
        row = 1
        ws_cost.cell(row=row + 1, column=1, value='Market Price per BTC')
        ws_cost.cell(row=row + 1, column=2, value=round(btc_price_usd, 2))
        ws_cost.cell(row=row + 1, column=3, value='USD')
        row += 1
        
        ws_cost.cell(row=row + 1, column=1, value='My Cost Basis per BTC')
        ws_cost.cell(row=row + 1, column=2, value=round(cost_basis_usd_per_btc, 2))
        ws_cost.cell(row=row + 1, column=3, value='USD')
        row += 1
        
        ws_cost.cell(row=row + 1, column=1, value='Discount vs Market')
        ws_cost.cell(row=row + 1, column=2, value=round(discount_vs_market_pct, 2))
        ws_cost.cell(row=row + 1, column=3, value='%')
        row += 1
        
        # Sheet 6: Energy Metrics
        ws_energy = wb.create_sheet(title='Energy Metrics')
        ws_energy.cell(row=1, column=1, value='Energy Metric')
        ws_energy.cell(row=1, column=2, value='Value')
        ws_energy.cell(row=1, column=3, value='Unit')
        
        row = 1
        ws_energy.cell(row=row + 1, column=1, value='Power Consumption')
        ws_energy.cell(row=row + 1, column=2, value=round(power_watts / 1000, 2))
        ws_energy.cell(row=row + 1, column=3, value='kW')
        row += 1
        
        ws_energy.cell(row=row + 1, column=1, value='Daily Energy Usage')
        ws_energy.cell(row=row + 1, column=2, value=round(daily_energy_kwh, 2))
        ws_energy.cell(row=row + 1, column=3, value='kWh')
        row += 1
        
        ws_energy.cell(row=row + 1, column=1, value='Electricity Price')
        ws_energy.cell(row=row + 1, column=2, value=round(price_per_kwh, 5))
        ws_energy.cell(row=row + 1, column=3, value='$/kWh')
        row += 1
        
        ws_energy.cell(row=row + 1, column=1, value='Mining Efficiency')
        ws_energy.cell(row=row + 1, column=2, value=round(efficiency_w_th, 2))
        ws_energy.cell(row=row + 1, column=3, value='W/TH')
        row += 1
        
        ws_energy.cell(row=row + 1, column=1, value='Energy to Mining Ratio')
        ws_energy.cell(row=row + 1, column=2, value=round(energy_cost_percentage, 2))
        ws_energy.cell(row=row + 1, column=3, value='%')
        row += 1
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    platform_suffix = f'_{selected_platform_name.replace(" ", "_")}' if selected_platform else ''
    response['Content-Disposition'] = f'attachment; filename="forecasting_dashboard{platform_suffix}_export.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response
