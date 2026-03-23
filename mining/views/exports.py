from django.http import HttpResponse
from openpyxl import Workbook
from datetime import datetime
from ..services import (
    get_capex_opex_data, get_income_data, get_overview_data,
    get_forecasting_data, resolve_selected_platform,
)
from ..models import RemoteMiningPlatform, Miner, Payout, Expense, TopUp


def export_capex_opex_data(request):
    """Export CAPEX/OPEX dashboard data to Excel file"""
    
    wb = Workbook()
    data = get_capex_opex_data()
    total_expenses = data['total_expenses']
    total_capex = data['total_capex']
    total_opex = data['total_opex']
    platform_expenses = data['platform_expenses']
    monthly_capex = data['monthly_capex']
    monthly_capex_by_platform = data['monthly_capex_by_platform']
    monthly_opex = data['monthly_opex']
    monthly_opex_by_platform = data['monthly_opex_by_platform']
    all_months = data['all_months']

    # Sheet 1: Total Expenses Summary
    ws_summary = wb.create_sheet(title='Total Expenses Summary')
    
    # Headers
    ws_summary.cell(row=1, column=1, value='Expense Type')
    ws_summary.cell(row=1, column=2, value='Amount (USD)')
    
    # Data rows
    ws_summary.cell(row=2, column=1, value='Total Expenses')
    ws_summary.cell(row=2, column=2, value=float(total_expenses))
    
    ws_summary.cell(row=3, column=1, value='Total CAPEX')
    ws_summary.cell(row=3, column=2, value=float(total_capex))
    
    ws_summary.cell(row=4, column=1, value='Total OPEX')
    ws_summary.cell(row=4, column=2, value=float(total_opex))
    
    # Sheet 2: Expenses by Platform
    if platform_expenses:
        ws_platform = wb.create_sheet(title='Expenses by Platform')
        
        # Headers
        ws_platform.cell(row=1, column=1, value='Platform')
        ws_platform.cell(row=1, column=2, value='Total Expenses (USD)')
        ws_platform.cell(row=1, column=3, value='CAPEX (USD)')
        ws_platform.cell(row=1, column=4, value='OPEX (USD)')
        
        # Data rows
        for row, item in enumerate(platform_expenses, start=1):
            ws_platform.cell(row=row + 1, column=1, value=item['platform'].name)
            ws_platform.cell(row=row + 1, column=2, value=float(item['total']))
            ws_platform.cell(row=row + 1, column=3, value=float(item['capex']))
            ws_platform.cell(row=row + 1, column=4, value=float(item['opex']))
    
    # Sheet 3: Monthly CAPEX
    if monthly_capex and all_months:
        ws_monthly_capex = wb.create_sheet(title='Monthly CAPEX')
        
        # Headers
        ws_monthly_capex.cell(row=1, column=1, value='Month')
        ws_monthly_capex.cell(row=1, column=2, value='Total CAPEX (USD)')
        
        # Platform headers
        col = 2
        platform_cols = {}
        for platform in monthly_capex_by_platform.keys():
            ws_monthly_capex.cell(row=1, column=col + 1, value=f'{platform.name} CAPEX (USD)')
            platform_cols[platform] = col
            col += 1
        
        # Data rows
        for row, month in enumerate(all_months, start=1):
            if month:
                ws_monthly_capex.cell(row=row + 1, column=1, value=month.strftime('%Y-%m'))
                
                # Total CAPEX for this month
                month_total = Decimal('0')
                for item in monthly_capex:
                    if item['month'] == month:
                        month_total = item['total']
                        break
                ws_monthly_capex.cell(row=row + 1, column=2, value=float(month_total))
                
                # Platform CAPEX for this month
                for platform, platform_data in monthly_capex_by_platform.items():
                    platform_month_total = Decimal('0')
                    for item in platform_data:
                        if item['month'] == month:
                            platform_month_total = item['total']
                            break
                    ws_monthly_capex.cell(row=row + 1, column=platform_cols[platform] + 1, value=float(platform_month_total))
    
    # Sheet 4: Monthly OPEX
    if monthly_opex and all_months:
        ws_monthly_opex = wb.create_sheet(title='Monthly OPEX')
        
        # Headers
        ws_monthly_opex.cell(row=1, column=1, value='Month')
        ws_monthly_opex.cell(row=1, column=2, value='Total OPEX (USD)')
        
        # Platform headers
        col = 2
        platform_cols = {}
        for platform in monthly_opex_by_platform.keys():
            ws_monthly_opex.cell(row=1, column=col + 1, value=f'{platform.name} OPEX (USD)')
            platform_cols[platform] = col
            col += 1
        
        # Data rows
        for row, month in enumerate(all_months, start=1):
            if month:
                ws_monthly_opex.cell(row=row + 1, column=1, value=month.strftime('%Y-%m'))
                
                # Total OPEX for this month
                month_total = Decimal('0')
                for item in monthly_opex:
                    if item['month'] == month:
                        month_total = item['total']
                        break
                ws_monthly_opex.cell(row=row + 1, column=2, value=float(month_total))
                
                # Platform OPEX for this month
                for platform, platform_data in monthly_opex_by_platform.items():
                    platform_month_total = Decimal('0')
                    for item in platform_data:
                        if item['month'] == month:
                            platform_month_total = item['total']
                            break
                    ws_monthly_opex.cell(row=row + 1, column=platform_cols[platform] + 1, value=float(platform_month_total))
    
    # Generate response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="capex_opex_dashboard_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response


# Income Dashboard View


def export_income_data(request):
    """Export Income dashboard data to Excel file"""
    
    wb = Workbook()
    data = get_income_data()
    current_btc_price = data['current_btc_price']
    total_income_btc = data['total_income_btc']
    total_income_usd_then = data['total_income_usd_then']
    total_income_usd_now = data['total_income_usd_now']
    platform_income = data['platform_income']
    monthly_income_btc = data['monthly_income_btc']
    monthly_income_by_platform = data['monthly_income_by_platform']
    all_months = data['all_months']

    # Sheet 1: Total Income Summary
    ws_summary = wb.create_sheet(title='Total Income Summary')
    
    # Headers
    ws_summary.cell(row=1, column=1, value='Income Type')
    ws_summary.cell(row=1, column=2, value='Amount')
    
    # Data rows
    ws_summary.cell(row=2, column=1, value='Total Income BTC')
    ws_summary.cell(row=2, column=2, value=float(total_income_btc))
    
    ws_summary.cell(row=3, column=1, value='Total Income USD (then)')
    ws_summary.cell(row=3, column=2, value=float(total_income_usd_then))
    
    ws_summary.cell(row=4, column=1, value='Total Income USD (now)')
    ws_summary.cell(row=4, column=2, value=float(total_income_usd_now))
    
    # Sheet 2: Income by Platform
    if platform_income:
        ws_platform = wb.create_sheet(title='Income by Platform')
        
        # Headers
        ws_platform.cell(row=1, column=1, value='Platform')
        ws_platform.cell(row=1, column=2, value='Total Income BTC')
        ws_platform.cell(row=1, column=3, value='Total Income USD (then)')
        ws_platform.cell(row=1, column=4, value='Total Income USD (now)')
        
        # Data rows
        for row, item in enumerate(platform_income, start=1):
            ws_platform.cell(row=row + 1, column=1, value=item['platform'].name)
            ws_platform.cell(row=row + 1, column=2, value=float(item['total_btc']))
            ws_platform.cell(row=row + 1, column=3, value=float(item['total_usd_then']))
            ws_platform.cell(row=row + 1, column=4, value=float(item['total_usd_now']))
    
    # Sheet 3: Monthly Income BTC
    if monthly_income_btc and all_months:
        ws_monthly_btc = wb.create_sheet(title='Monthly Income BTC')
        
        # Headers
        ws_monthly_btc.cell(row=1, column=1, value='Month')
        ws_monthly_btc.cell(row=1, column=2, value='Total Income BTC')
        
        # Platform headers
        col = 2
        platform_cols = {}
        for platform in monthly_income_by_platform.keys():
            ws_monthly_btc.cell(row=1, column=col + 1, value=f'{platform.name} BTC')
            platform_cols[platform] = col
            col += 1
        
        # Data rows
        for row, month in enumerate(all_months, start=1):
            if month:
                ws_monthly_btc.cell(row=row + 1, column=1, value=month.strftime('%Y-%m'))
                
                # Total BTC for this month
                month_total = Decimal('0')
                for item in monthly_income_btc:
                    if item['month'] == month:
                        month_total = item['total_btc']
                        break
                ws_monthly_btc.cell(row=row + 1, column=2, value=float(month_total))
                
                # Platform BTC for this month
                for platform, platform_data in monthly_income_by_platform.items():
                    platform_month_total = Decimal('0')
                    for item in platform_data:
                        if item['month'] == month:
                            platform_month_total = item['total_btc']
                            break
                    ws_monthly_btc.cell(row=row + 1, column=platform_cols[platform] + 1, value=float(platform_month_total))
    
    # Sheet 4: Monthly Income USD (then)
    if monthly_income_btc and all_months:
        ws_monthly_usd_then = wb.create_sheet(title='Monthly Income USD then')
        
        # Headers
        ws_monthly_usd_then.cell(row=1, column=1, value='Month')
        ws_monthly_usd_then.cell(row=1, column=2, value='Total Income USD (then)')
        
        # Platform headers
        col = 2
        platform_cols = {}
        for platform in monthly_income_by_platform.keys():
            ws_monthly_usd_then.cell(row=1, column=col + 1, value=f'{platform.name} USD (then)')
            platform_cols[platform] = col
            col += 1
        
        # Data rows
        for row, month in enumerate(all_months, start=1):
            if month:
                ws_monthly_usd_then.cell(row=row + 1, column=1, value=month.strftime('%Y-%m'))
                
                # Total USD then for this month
                month_total = Decimal('0')
                for item in monthly_income_btc:
                    if item['month'] == month:
                        month_total = item['total_usd_then']
                        break
                ws_monthly_usd_then.cell(row=row + 1, column=2, value=float(month_total))
                
                # Platform USD then for this month
                for platform, platform_data in monthly_income_by_platform.items():
                    platform_month_total = Decimal('0')
                    for item in platform_data:
                        if item['month'] == month:
                            platform_month_total = item['total_usd_then']
                            break
                    ws_monthly_usd_then.cell(row=row + 1, column=platform_cols[platform] + 1, value=float(platform_month_total))
    
    # Sheet 5: Monthly Income USD (now)
    if monthly_income_btc and all_months:
        ws_monthly_usd_now = wb.create_sheet(title='Monthly Income USD now')
        
        # Headers
        ws_monthly_usd_now.cell(row=1, column=1, value='Month')
        ws_monthly_usd_now.cell(row=1, column=2, value='Total Income USD (now)')
        
        # Platform headers
        col = 2
        platform_cols = {}
        for platform in monthly_income_by_platform.keys():
            ws_monthly_usd_now.cell(row=1, column=col + 1, value=f'{platform.name} USD (now)')
            platform_cols[platform] = col
            col += 1
        
        # Data rows
        for row, month in enumerate(all_months, start=1):
            if month:
                ws_monthly_usd_now.cell(row=row + 1, column=1, value=month.strftime('%Y-%m'))
                
                # Total USD now for this month
                month_total = Decimal('0')
                for item in monthly_income_btc:
                    if item['month'] == month:
                        month_total = item['total_usd_now']
                        break
                ws_monthly_usd_now.cell(row=row + 1, column=2, value=float(month_total))
                
                # Platform USD now for this month
                for platform, platform_data in monthly_income_by_platform.items():
                    platform_month_total = Decimal('0')
                    for item in platform_data:
                        if item['month'] == month:
                            platform_month_total = item['total_usd_now']
                            break
                    ws_monthly_usd_now.cell(row=row + 1, column=platform_cols[platform] + 1, value=float(platform_month_total))
    
    # Generate response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="income_dashboard_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response




def download_platform_template(request):
    """Download import template for Remote Mining Platforms"""
    wb = Workbook()
    ws = wb.create_sheet(title='Platform Import Template')
    
    # Add headers based on form fields
    headers = ['name', 'website_link', 'portal_url', 'point_of_contact_name', 
               'point_of_contact_email', 'point_of_contact_phone', 'point_of_contact_telegram', 'energy_price']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="platform_import_template.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response




def download_miner_template(request):
    """Download import template for Miners"""
    wb = Workbook()
    ws = wb.create_sheet(title='Miner Import Template')
    
    # Add headers based on form fields (excluding image field for import)
    headers = ['model', 'manufacturer', 'product_link', 'serial_number', 
               'platform', 'platform_internal_id', 'hashrate', 'power', 'efficiency', 
               'purchase_price', 'purchase_date', 'start_date', 'location']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="miner_import_template.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response




def download_payout_template(request):
    """Download import template for Payouts"""
    wb = Workbook()
    ws = wb.create_sheet(title='Payout Import Template')
    
    # Add headers based on form fields
    headers = ['payout_date', 'payout_amount', 'platform', 'transaction_id', 'closing_price', 'value_at_payout (read-only)']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="payout_import_template.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response


# Data Export Views


def export_platform_data(request):
    """Export all platform data to Excel file"""
    wb = Workbook()
    ws = wb.create_sheet(title='Platform Data')
    
    # Add headers
    headers = ['name', 'website_link', 'portal_url', 'point_of_contact_name', 
               'point_of_contact_email', 'point_of_contact_phone', 
               'point_of_contact_telegram', 'energy_price']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    # Add data rows
    platforms = RemoteMiningPlatform.objects.all()
    for row, platform in enumerate(platforms, start=1):
        ws.cell(row=row + 1, column=1, value=platform.name or '')
        ws.cell(row=row + 1, column=2, value=platform.website_link or '')
        ws.cell(row=row + 1, column=3, value=platform.portal_url or '')
        ws.cell(row=row + 1, column=4, value=platform.point_of_contact_name or '')
        ws.cell(row=row + 1, column=5, value=platform.point_of_contact_email or '')
        ws.cell(row=row + 1, column=6, value=platform.point_of_contact_phone or '')
        ws.cell(row=row + 1, column=7, value=platform.point_of_contact_telegram or '')
        ws.cell(row=row + 1, column=8, value=float(platform.energy_price) if platform.energy_price else '')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="platform_data_export.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response




def export_miner_data(request):
    """Export all miner data to Excel file"""
    wb = Workbook()
    ws = wb.create_sheet(title='Miner Data')
    
    # Add headers
    headers = ['model', 'manufacturer', 'product_link', 'serial_number', 
               'platform', 'platform_internal_id', 'hashrate', 'power', 'efficiency', 
               'purchase_price', 'purchase_date', 'start_date', 'location']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    # Add data rows
    miners = Miner.objects.select_related('platform').all()
    for row, miner in enumerate(miners, start=1):
        ws.cell(row=row + 1, column=1, value=miner.model or '')
        ws.cell(row=row + 1, column=2, value=miner.manufacturer or '')
        ws.cell(row=row + 1, column=3, value=miner.product_link or '')
        ws.cell(row=row + 1, column=4, value=miner.serial_number or '')
        ws.cell(row=row + 1, column=5, value=miner.platform.pk if miner.platform else '')
        ws.cell(row=row + 1, column=6, value=miner.platform_internal_id or '')
        ws.cell(row=row + 1, column=7, value=float(miner.hashrate) if miner.hashrate else '')
        ws.cell(row=row + 1, column=8, value=float(miner.power) if miner.power else '')
        ws.cell(row=row + 1, column=9, value=float(miner.efficiency) if miner.efficiency else '')
        ws.cell(row=row + 1, column=10, value=float(miner.purchase_price) if miner.purchase_price else '')
        ws.cell(row=row + 1, column=11, value=miner.purchase_date.strftime('%Y-%m-%d') if miner.purchase_date else '')
        ws.cell(row=row + 1, column=12, value=miner.start_date.strftime('%Y-%m-%d') if miner.start_date else '')
        ws.cell(row=row + 1, column=13, value=miner.location or '')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="miner_data_export.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response




def export_payout_data(request):
    """Export all payout data to Excel file"""
    wb = Workbook()
    ws = wb.create_sheet(title='Payout Data')
    
    # Add headers
    headers = ['payout_date', 'payout_amount', 'platform', 'transaction_id', 'closing_price', 'value_at_payout']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    # Add data rows
    payouts = Payout.objects.select_related('platform').all()
    for row, payout in enumerate(payouts, start=1):
        ws.cell(row=row + 1, column=1, value=payout.payout_date.strftime('%Y-%m-%d'))
        ws.cell(row=row + 1, column=2, value=float(payout.payout_amount))
        ws.cell(row=row + 1, column=3, value=payout.platform.pk if payout.platform else '')
        ws.cell(row=row + 1, column=4, value=payout.transaction_id or '')
        ws.cell(row=row + 1, column=5, value=float(payout.closing_price) if payout.closing_price else '')
        ws.cell(row=row + 1, column=6, value=float(payout.value_at_payout) if payout.value_at_payout else '')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="payout_data_export.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response




def export_overview_data(request):
    """Export overview dashboard data to Excel file"""
    wb = Workbook()
    selected_platform = resolve_selected_platform(request.GET.get('platform', ''))
    data = get_overview_data(selected_platform)
    selected_platform_name = selected_platform.name if selected_platform else 'All Platforms'
    bitcoin_price = data['bitcoin_price']
    network_hashrate = data['network_hashrate']
    network_difficulty = data['network_difficulty']
    avg_block_fees_24h = data['avg_block_fees_24h']
    miner_count = data['miner_count']
    total_hashrate = data['total_hashrate']
    total_power = data['total_power']
    total_capex = data['total_capex']
    avg_efficiency = data['avg_efficiency']
    hashrate_weighted_efficiency = data['hashrate_weighted_efficiency']
    avg_energy_cost = data['avg_energy_cost']
    hashrate_weighted_energy_cost = data['hashrate_weighted_energy_cost']
    total_btc_mined = data['total_btc_mined']
    current_gross_value = data['current_gross_value']
    gross_value_at_payout = data['gross_value_at_payout']
    appreciation = data['appreciation']
    total_opex = data['total_opex']
    current_net_value = data['current_net_value']
    total_payouts = data['total_payouts']

    # Sheet 1: Overview Summary
    ws_summary = wb.create_sheet(title='Overview Summary')
    
    # Headers
    ws_summary.cell(row=1, column=1, value='Metric Category')
    ws_summary.cell(row=1, column=2, value='Metric Name')
    ws_summary.cell(row=1, column=3, value='Value')
    ws_summary.cell(row=1, column=4, value='Unit')
    
    # Data rows
    row = 1
    
    # Platform Filter
    ws_summary.cell(row=row + 1, column=1, value='Filter')
    ws_summary.cell(row=row + 1, column=2, value='Platform')
    ws_summary.cell(row=row + 1, column=3, value=selected_platform_name)
    ws_summary.cell(row=row + 1, column=4, value='')
    row += 1
    
    # Network Data
    ws_summary.cell(row=row + 1, column=1, value='Network Data')
    ws_summary.cell(row=row + 1, column=2, value='Bitcoin Spot Price')
    ws_summary.cell(row=row + 1, column=3, value=float(bitcoin_price))
    ws_summary.cell(row=row + 1, column=4, value='USD')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Network Data')
    ws_summary.cell(row=row + 1, column=2, value='Total Network Hashrate')
    ws_summary.cell(row=row + 1, column=3, value=float(network_hashrate))
    ws_summary.cell(row=row + 1, column=4, value='EH/s')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Network Data')
    ws_summary.cell(row=row + 1, column=2, value='Network Difficulty')
    ws_summary.cell(row=row + 1, column=3, value=float(network_difficulty))
    ws_summary.cell(row=row + 1, column=4, value='')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Network Data')
    ws_summary.cell(row=row + 1, column=2, value='24h Avg Block Fees')
    ws_summary.cell(row=row + 1, column=3, value=float(avg_block_fees_24h))
    ws_summary.cell(row=row + 1, column=4, value='BTC')
    row += 1
    
    # Fleet Data
    ws_summary.cell(row=row + 1, column=1, value='Fleet Data')
    ws_summary.cell(row=row + 1, column=2, value='Miner Count')
    ws_summary.cell(row=row + 1, column=3, value=miner_count)
    ws_summary.cell(row=row + 1, column=4, value='units')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Fleet Data')
    ws_summary.cell(row=row + 1, column=2, value='Total Hashrate')
    ws_summary.cell(row=row + 1, column=3, value=float(total_hashrate))
    ws_summary.cell(row=row + 1, column=4, value='TH/s')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Fleet Data')
    ws_summary.cell(row=row + 1, column=2, value='Total Power')
    ws_summary.write(row, 2, round(float(total_power), 2))  # Power already stored in kW in database
    ws_summary.cell(row=row + 1, column=4, value='kW')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Fleet Data')
    ws_summary.cell(row=row + 1, column=2, value='Total Hardware Cost')
    ws_summary.cell(row=row + 1, column=3, value=float(total_capex))
    ws_summary.cell(row=row + 1, column=4, value='USD')
    row += 1
    
    # Efficiency Data
    ws_summary.cell(row=row + 1, column=1, value='Efficiency Data')
    ws_summary.cell(row=row + 1, column=2, value='Average Efficiency')
    ws_summary.cell(row=row + 1, column=3, value=float(avg_efficiency))
    ws_summary.cell(row=row + 1, column=4, value='W/TH')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Efficiency Data')
    ws_summary.cell(row=row + 1, column=2, value='Hashrate Weighted Avg Efficiency')
    ws_summary.cell(row=row + 1, column=3, value=float(hashrate_weighted_efficiency))
    ws_summary.cell(row=row + 1, column=4, value='W/TH')
    row += 1
    
    # Energy Data
    ws_summary.cell(row=row + 1, column=1, value='Energy Data')
    ws_summary.cell(row=row + 1, column=2, value='Average Energy Cost')
    ws_summary.cell(row=row + 1, column=3, value=float(avg_energy_cost))
    ws_summary.cell(row=row + 1, column=4, value='$/kWh')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Energy Data')
    ws_summary.cell(row=row + 1, column=2, value='Hashrate Weighted Avg Energy Cost')
    ws_summary.cell(row=row + 1, column=3, value=float(hashrate_weighted_energy_cost))
    ws_summary.cell(row=row + 1, column=4, value='$/kWh')
    row += 1
    
    # Revenue Data
    ws_summary.cell(row=row + 1, column=1, value='Revenue Data')
    ws_summary.cell(row=row + 1, column=2, value='Total BTC Mined')
    ws_summary.cell(row=row + 1, column=3, value=float(total_btc_mined))
    ws_summary.cell(row=row + 1, column=4, value='BTC')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Revenue Data')
    ws_summary.cell(row=row + 1, column=2, value='Current Gross Value')
    ws_summary.cell(row=row + 1, column=3, value=round(float(current_gross_value), 2))
    ws_summary.cell(row=row + 1, column=4, value='USD')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Revenue Data')
    ws_summary.cell(row=row + 1, column=2, value='Gross Value at Payout')
    ws_summary.cell(row=row + 1, column=3, value=round(float(gross_value_at_payout), 2))
    ws_summary.cell(row=row + 1, column=4, value='USD')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Revenue Data')
    ws_summary.cell(row=row + 1, column=2, value='Appreciation')
    ws_summary.cell(row=row + 1, column=3, value=round(float(appreciation), 2))
    ws_summary.cell(row=row + 1, column=4, value='USD')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Revenue Data')
    ws_summary.cell(row=row + 1, column=2, value='Total Payouts')
    ws_summary.cell(row=row + 1, column=3, value=total_payouts)
    ws_summary.cell(row=row + 1, column=4, value='count')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Revenue Data')
    ws_summary.cell(row=row + 1, column=2, value='Total OPEX')
    ws_summary.cell(row=row + 1, column=3, value=round(float(total_opex), 2))
    ws_summary.cell(row=row + 1, column=4, value='USD')
    row += 1
    
    ws_summary.cell(row=row + 1, column=1, value='Revenue Data')
    ws_summary.cell(row=row + 1, column=2, value='Current Net Value')
    ws_summary.cell(row=row + 1, column=3, value=round(float(current_net_value), 2))
    ws_summary.cell(row=row + 1, column=4, value='USD')
    row += 1
    
    # Sheet 2: Hashrate by Platform
    ws_hashrate_platform = wb.create_sheet(title='Hashrate by Platform')
    ws_hashrate_platform.cell(row=1, column=1, value='Platform')
    ws_hashrate_platform.cell(row=1, column=2, value='Hashrate (TH/s)')

    platform_row = 1
    for item in data['hashrate_by_platform']:
        ws_hashrate_platform.cell(row=platform_row + 1, column=1, value=item['platform'])
        ws_hashrate_platform.cell(row=platform_row + 1, column=2, value=item['hashrate'])
        platform_row += 1

    # Sheet 3: Hashrate by Location
    ws_hashrate_location = wb.create_sheet(title='Hashrate by Location')
    ws_hashrate_location.cell(row=1, column=1, value='Location')
    ws_hashrate_location.cell(row=1, column=2, value='Hashrate (TH/s)')

    location_row = 1
    for item in data['hashrate_by_location']:
        ws_hashrate_location.cell(row=location_row + 1, column=1, value=item['location'])
        ws_hashrate_location.cell(row=location_row + 1, column=2, value=item['hashrate'])
        location_row += 1

    # Sheet 4: Revenue by Platform
    ws_revenue_platform = wb.create_sheet(title='Revenue by Platform')
    ws_revenue_platform.cell(row=1, column=1, value='Platform')
    ws_revenue_platform.cell(row=1, column=2, value='BTC Mined')
    ws_revenue_platform.cell(row=1, column=3, value='Gross Value (USD)')
    ws_revenue_platform.cell(row=1, column=4, value='Gross Value at Payout (USD)')
    ws_revenue_platform.cell(row=1, column=5, value='Payout Count')

    revenue_row = 1
    for item in data['revenue_by_platform']:
        ws_revenue_platform.cell(row=revenue_row + 1, column=1, value=item['platform'])
        ws_revenue_platform.cell(row=revenue_row + 1, column=2, value=item['btc_mined'])
        ws_revenue_platform.cell(row=revenue_row + 1, column=3, value=round(item['gross_value'], 2))
        ws_revenue_platform.cell(row=revenue_row + 1, column=4, value=round(item['gross_value_at_payout'], 2))
        ws_revenue_platform.cell(row=revenue_row + 1, column=5, value=item['payout_count'])
        revenue_row += 1
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    platform_suffix = f'_{selected_platform_name.replace(" ", "_")}' if selected_platform else ''
    response['Content-Disposition'] = f'attachment; filename="overview_dashboard{platform_suffix}_export.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response


# Data Import Views


def download_expense_template(request):
    """Download import template for Expenses"""
    wb = Workbook()
    ws = wb.create_sheet(title='Expense Import Template')
    
    # Add headers based on form fields
    headers = ['expense_date', 'platform', 'category', 'description', 'expense_amount', 'invoice_link', 'receipt_link', 'notes']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="expense_import_template.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response




def export_expense_data(request):
    """Export all expense data to Excel file"""
    wb = Workbook()
    ws = wb.create_sheet(title='Expense Data')
    
    # Add headers
    headers = ['expense_date', 'platform', 'category', 'description', 'expense_amount', 'invoice_link', 'receipt_link', 'notes']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    # Add data rows
    expenses = Expense.objects.select_related('platform').all().order_by('-expense_date')
    for row, expense in enumerate(expenses, start=1):
        ws.cell(row=row + 1, column=1, value=expense.expense_date.strftime('%Y-%m-%d') if expense.expense_date else '')
        ws.cell(row=row + 1, column=2, value=expense.platform.id if expense.platform else '')
        ws.cell(row=row + 1, column=3, value=expense.category or '')
        ws.cell(row=row + 1, column=4, value=expense.description or '')
        ws.cell(row=row + 1, column=5, value=float(expense.expense_amount) if expense.expense_amount else '')
        ws.cell(row=row + 1, column=6, value=expense.invoice_link or '')
        ws.cell(row=row + 1, column=7, value=expense.receipt_link or '')
        ws.cell(row=row + 1, column=8, value=expense.notes or '')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="expense_data_export.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response




def download_topup_template(request):
    """Download Excel template for Top-Up import"""
    wb = Workbook()
    ws = wb.create_sheet(title='TopUp Template')
    
    # Add headers
    headers = ['topup_date', 'platform', 'topup_amount', 'description', 'receipt_link']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="topup_import_template.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response




def export_topup_data(request):
    """Export all top-up data to Excel file"""
    wb = Workbook()
    ws = wb.create_sheet(title='TopUp Data')
    
    # Add headers
    headers = ['topup_date', 'platform', 'topup_amount', 'description', 'receipt_link']
    
    for col, header in enumerate(headers):
        ws.cell(row=1, column=col + 1, value=header)
    
    # Add data rows
    topups = TopUp.objects.select_related('platform').all().order_by('-topup_date')
    for row, topup in enumerate(topups, start=1):
        ws.cell(row=row + 1, column=1, value=str(topup.topup_date) if topup.topup_date else '')
        ws.cell(row=row + 1, column=2, value=topup.platform.name if topup.platform else '')
        ws.cell(row=row + 1, column=3, value=float(topup.topup_amount) if topup.topup_amount else '')
        ws.cell(row=row + 1, column=4, value=topup.description or '')
        ws.cell(row=row + 1, column=5, value=topup.receipt_link or '')
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="topup_data_export.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response




def export_forecasting_data(request):
    """Export forecasting dashboard data to Excel file"""
    wb = Workbook()
    selected_platform = resolve_selected_platform(request.GET.get('platform', ''))
    data = get_forecasting_data(selected_platform)
    selected_platform_name = selected_platform.name if selected_platform else 'All Platforms'
    total_hashrate = data['total_hashrate']
    miner_count = data['miner_count']
    total_miner_count = data['total_miner_count']
    network_difficulty = data['network_difficulty']
    network_hashrate_ehs = data['network_hashrate_ehs']
    avg_tx_fees = data['avg_tx_fees']
    pool_fee = data['pool_fee']
    btc_price_usd = data['btc_price_usd']
    avg_efficiency = data['avg_efficiency']
    hashrate_weighted_efficiency = data['hashrate_weighted_efficiency']
    avg_energy_cost = data['avg_energy_cost']
    hashrate_weighted_energy_cost = data['hashrate_weighted_energy_cost']
    results = data['results']

    # Check if we have required data
    if results is None:
        ws = wb.create_sheet(title='Error')
        ws.cell(row=1, column=1, value='Error: Insufficient data (missing miners, API data, or network difficulty)')
    else:
        power_watts = results['power_consumption_watts']
        margin = results['margin']
        energy_cost_percentage = results['energy_cost_percentage']
        hashrate_share_percent = results['hashrate_share_percent']
        days_to_mine_1_btc = results['days_to_mine_1_btc']
        days_to_mine_small_btc = results['days_to_mine_small_btc']
        roi_data = results['roi_data']

        def format_time_breakdown(days_total):
            if days_total == float('inf'):
                return 'Never (not profitable)'
            years = int(days_total / 365)
            months = int((days_total % 365) / 30)
            remaining_days = int(days_total % 30)
            return f"{years} years, {months} months, {remaining_days} days"

        time_to_mine_1_btc = format_time_breakdown(days_to_mine_1_btc)
        time_to_mine_small_btc = format_time_breakdown(days_to_mine_small_btc)

        # Sheet 1: Forecasting Summary
        ws_summary = wb.create_sheet(title='Forecasting Summary')
        ws_summary.cell(row=1, column=1, value='Metric Category')
        ws_summary.cell(row=1, column=2, value='Metric Name')
        ws_summary.cell(row=1, column=3, value='Value')
        ws_summary.cell(row=1, column=4, value='Unit')
        
        row = 1
        
        # Platform Filter
        ws_summary.cell(row=row + 1, column=1, value='Filter')
        ws_summary.cell(row=row + 1, column=2, value='Platform')
        ws_summary.cell(row=row + 1, column=3, value=selected_platform_name)
        ws_summary.cell(row=row + 1, column=4, value='')
        row += 1
        
        # Network Overview
        ws_summary.cell(row=row + 1, column=1, value='Network Overview')
        ws_summary.cell(row=row + 1, column=2, value='Network Difficulty')
        ws_summary.cell(row=row + 1, column=3, value=network_difficulty)
        ws_summary.cell(row=row + 1, column=4, value='')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='Network Overview')
        ws_summary.cell(row=row + 1, column=2, value='Network Hashrate')
        ws_summary.cell(row=row + 1, column=3, value=network_hashrate_ehs)
        ws_summary.cell(row=row + 1, column=4, value='EH/s')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='Network Overview')
        ws_summary.cell(row=row + 1, column=2, value='Avg Block Fees (24h)')
        ws_summary.cell(row=row + 1, column=3, value=round(avg_tx_fees, 8))
        ws_summary.cell(row=row + 1, column=4, value='BTC')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='Network Overview')
        ws_summary.cell(row=row + 1, column=2, value='BTC Price')
        ws_summary.cell(row=row + 1, column=3, value=round(btc_price_usd, 2))
        ws_summary.cell(row=row + 1, column=4, value='USD')
        row += 1
        
        # My Fleet Overview
        ws_summary.cell(row=row + 1, column=1, value='My Fleet Overview')
        ws_summary.cell(row=row + 1, column=2, value='Miners Accounted For')
        ws_summary.cell(row=row + 1, column=3, value=miner_count)
        ws_summary.cell(row=row + 1, column=4, value=f'of {total_miner_count}')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='My Fleet Overview')
        ws_summary.cell(row=row + 1, column=2, value='My Hashrate')
        ws_summary.cell(row=row + 1, column=3, value=float(total_hashrate))
        ws_summary.cell(row=row + 1, column=4, value='TH/s')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='My Fleet Overview')
        ws_summary.cell(row=row + 1, column=2, value='Pool Fee')
        ws_summary.cell(row=row + 1, column=3, value=pool_fee)
        ws_summary.cell(row=row + 1, column=4, value='%')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='My Fleet Overview')
        ws_summary.cell(row=row + 1, column=2, value='Power Consumption')
        ws_summary.cell(row=row + 1, column=3, value=round(power_watts / 1000, 2))
        ws_summary.cell(row=row + 1, column=4, value='kW')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='My Fleet Overview')
        ws_summary.cell(row=row + 1, column=2, value='Net Profit Margin')
        ws_summary.cell(row=row + 1, column=3, value=round(margin, 2))
        ws_summary.cell(row=row + 1, column=4, value='%')
        row += 1
        
        # Efficiency Data
        ws_summary.cell(row=row + 1, column=1, value='Efficiency Data')
        ws_summary.cell(row=row + 1, column=2, value='Average Efficiency')
        ws_summary.cell(row=row + 1, column=3, value=float(avg_efficiency))
        ws_summary.cell(row=row + 1, column=4, value='W/TH')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='Efficiency Data')
        ws_summary.cell(row=row + 1, column=2, value='Hashrate Weighted Avg Efficiency')
        ws_summary.cell(row=row + 1, column=3, value=round(float(hashrate_weighted_efficiency), 2))
        ws_summary.cell(row=row + 1, column=4, value='W/TH')
        row += 1
        
        # Energy Data
        ws_summary.cell(row=row + 1, column=1, value='Energy Data')
        ws_summary.cell(row=row + 1, column=2, value='Average Energy Cost')
        ws_summary.cell(row=row + 1, column=3, value=float(avg_energy_cost))
        ws_summary.cell(row=row + 1, column=4, value='$/kWh')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='Energy Data')
        ws_summary.cell(row=row + 1, column=2, value='Hashrate Weighted Avg Energy Cost')
        ws_summary.cell(row=row + 1, column=3, value=round(float(hashrate_weighted_energy_cost), 6))
        ws_summary.cell(row=row + 1, column=4, value='$/kWh')
        row += 1
        
        # Key Metrics
        ws_summary.cell(row=row + 1, column=1, value='Key Metrics')
        ws_summary.cell(row=row + 1, column=2, value='Time to mine 1 BTC')
        ws_summary.cell(row=row + 1, column=3, value=time_to_mine_1_btc)
        ws_summary.cell(row=row + 1, column=4, value='')
        row += 1
        
        ws_summary.cell(row=row + 1, column=1, value='Key Metrics')
        ws_summary.cell(row=row + 1, column=2, value='Time to mine 0.005 BTC')
        ws_summary.cell(row=row + 1, column=3, value=time_to_mine_small_btc)
        ws_summary.cell(row=row + 1, column=4, value='')
        row += 1
        
        if roi_data:
            roi_display = f"{roi_data['time_breakdown']['years']} years, {roi_data['time_breakdown']['months']} months, {roi_data['time_breakdown']['days']} days" if roi_data['days_to_roi'] != float('inf') else "Never (not profitable)"
            ws_summary.cell(row=row + 1, column=1, value='Key Metrics')
            ws_summary.cell(row=row + 1, column=2, value='ROI')
            ws_summary.cell(row=row + 1, column=3, value=roi_display)
            ws_summary.cell(row=row + 1, column=4, value='')
            row += 1
        
        # Sheet 2: Daily Projections
        ws_daily = wb.create_sheet(title='Daily Projections')
        ws_daily.cell(row=1, column=1, value='Projection Type')
        ws_daily.cell(row=1, column=2, value='BTC Amount')
        ws_daily.cell(row=1, column=3, value='USD Amount')
        
        row = 1
        ws_daily.cell(row=row + 1, column=1, value='Gross Payout (before pool fee)')
        ws_daily.cell(row=row + 1, column=2, value=round(daily_btc_gross_before_fee, 8))
        ws_daily.cell(row=row + 1, column=3, value=round(daily_usd_gross, 2))
        row += 1
        
        ws_daily.cell(row=row + 1, column=1, value='Pool Fee')
        ws_daily.cell(row=row + 1, column=2, value=round(pool_fee_btc, 8))
        ws_daily.cell(row=row + 1, column=3, value=round(pool_fee_btc * btc_price_usd, 2))
        row += 1
        
        ws_daily.cell(row=row + 1, column=1, value='After Pool Fee')
        ws_daily.cell(row=row + 1, column=2, value=round(daily_btc_after_fee, 8))
        ws_daily.cell(row=row + 1, column=3, value=round(daily_usd_after_fee, 2))
        row += 1
        
        ws_daily.cell(row=row + 1, column=1, value='Electricity Cost')
        ws_daily.cell(row=row + 1, column=2, value=round(daily_electricity_cost_usd / btc_price_usd, 8) if btc_price_usd > 0 else 0)
        ws_daily.cell(row=row + 1, column=3, value=round(daily_electricity_cost_usd, 2))
        row += 1
        
        ws_daily.cell(row=row + 1, column=1, value='Net Profit')
        ws_daily.cell(row=row + 1, column=2, value=round(daily_btc_net, 8))
        ws_daily.cell(row=row + 1, column=3, value=round(daily_usd_net, 2))
        row += 1
        
        # Sheet 3: Monthly Projections
        ws_monthly = wb.create_sheet(title='Monthly Projections')
        ws_monthly.cell(row=1, column=1, value='Projection Type')
        ws_monthly.cell(row=1, column=2, value='BTC Amount')
        ws_monthly.cell(row=1, column=3, value='USD Amount')
        
        row = 1
        ws_monthly.cell(row=row + 1, column=1, value='Gross Payout (before pool fee)')
        ws_monthly.cell(row=row + 1, column=2, value=round(daily_btc_gross_before_fee * 30, 8))
        ws_monthly.cell(row=row + 1, column=3, value=round(daily_usd_gross * 30, 2))
        row += 1
        
        ws_monthly.cell(row=row + 1, column=1, value='Pool Fee')
        ws_monthly.cell(row=row + 1, column=2, value=round(pool_fee_btc * 30, 8))
        ws_monthly.cell(row=row + 1, column=3, value=round(pool_fee_btc * btc_price_usd * 30, 2))
        row += 1
        
        ws_monthly.cell(row=row + 1, column=1, value='After Pool Fee')
        ws_monthly.cell(row=row + 1, column=2, value=round(daily_btc_after_fee * 30, 8))
        ws_monthly.cell(row=row + 1, column=3, value=round(daily_usd_after_fee * 30, 2))
        row += 1
        
        ws_monthly.cell(row=row + 1, column=1, value='Electricity Cost')
        ws_monthly.cell(row=row + 1, column=2, value=round(daily_electricity_cost_usd / btc_price_usd * 30, 8) if btc_price_usd > 0 else 0)
        ws_monthly.cell(row=row + 1, column=3, value=round(daily_electricity_cost_usd * 30, 2))
        row += 1
        
        ws_monthly.cell(row=row + 1, column=1, value='Net Profit')
        ws_monthly.cell(row=row + 1, column=2, value=round(daily_btc_net * 30, 8))
        ws_monthly.cell(row=row + 1, column=3, value=round(daily_usd_net * 30, 2))
        row += 1
        
        # Sheet 4: Yearly Projections
        ws_yearly = wb.create_sheet(title='Yearly Projections')
        ws_yearly.cell(row=1, column=1, value='Projection Type')
        ws_yearly.cell(row=1, column=2, value='BTC Amount')
        ws_yearly.cell(row=1, column=3, value='USD Amount')
        
        row = 1
        ws_yearly.cell(row=row + 1, column=1, value='Gross Payout (before pool fee)')
        ws_yearly.cell(row=row + 1, column=2, value=round(daily_btc_gross_before_fee * 365, 8))
        ws_yearly.cell(row=row + 1, column=3, value=round(daily_usd_gross * 365, 2))
        row += 1
        
        ws_yearly.cell(row=row + 1, column=1, value='Pool Fee')
        ws_yearly.cell(row=row + 1, column=2, value=round(pool_fee_btc * 365, 8))
        ws_yearly.cell(row=row + 1, column=3, value=round(pool_fee_btc * btc_price_usd * 365, 2))
        row += 1
        
        ws_yearly.cell(row=row + 1, column=1, value='After Pool Fee')
        ws_yearly.cell(row=row + 1, column=2, value=round(daily_btc_after_fee * 365, 8))
        ws_yearly.cell(row=row + 1, column=3, value=round(daily_usd_after_fee * 365, 2))
        row += 1
        
        ws_yearly.cell(row=row + 1, column=1, value='Electricity Cost')
        ws_yearly.cell(row=row + 1, column=2, value=round(daily_electricity_cost_usd / btc_price_usd * 365, 8) if btc_price_usd > 0 else 0)
        ws_yearly.cell(row=row + 1, column=3, value=round(daily_electricity_cost_usd * 365, 2))
        row += 1
        
        ws_yearly.cell(row=row + 1, column=1, value='Net Profit')
        ws_yearly.cell(row=row + 1, column=2, value=round(daily_btc_net * 365, 8))
        ws_yearly.cell(row=row + 1, column=3, value=round(daily_usd_net * 365, 2))
        row += 1
        
        # Sheet 5: Cost Basis Analysis
        ws_cost = wb.create_sheet(title='Cost Basis Analysis')
        ws_cost.cell(row=1, column=1, value='Cost Analysis')
        ws_cost.cell(row=1, column=2, value='Value')
        ws_cost.cell(row=1, column=3, value='Unit')
        
        row = 1
        ws_cost.cell(row=row + 1, column=1, value='Market Price per BTC')
        ws_cost.cell(row=row + 1, column=2, value=round(btc_price_usd, 2))
        ws_cost.cell(row=row + 1, column=3, value='USD')
        row += 1
        
        ws_cost.cell(row=row + 1, column=1, value='My Cost Basis per BTC')
        ws_cost.cell(row=row + 1, column=2, value=round(cost_basis_usd_per_btc, 2))
        ws_cost.cell(row=row + 1, column=3, value='USD')
        row += 1
        
        ws_cost.cell(row=row + 1, column=1, value='Discount vs Market')
        ws_cost.cell(row=row + 1, column=2, value=round(discount_vs_market_pct, 2))
        ws_cost.cell(row=row + 1, column=3, value='%')
        row += 1
        
        # Sheet 6: Energy Metrics
        ws_energy = wb.create_sheet(title='Energy Metrics')
        ws_energy.cell(row=1, column=1, value='Energy Metric')
        ws_energy.cell(row=1, column=2, value='Value')
        ws_energy.cell(row=1, column=3, value='Unit')
        
        row = 1
        ws_energy.cell(row=row + 1, column=1, value='Power Consumption')
        ws_energy.cell(row=row + 1, column=2, value=round(power_watts / 1000, 2))
        ws_energy.cell(row=row + 1, column=3, value='kW')
        row += 1
        
        ws_energy.cell(row=row + 1, column=1, value='Daily Energy Usage')
        ws_energy.cell(row=row + 1, column=2, value=round(daily_energy_kwh, 2))
        ws_energy.cell(row=row + 1, column=3, value='kWh')
        row += 1
        
        ws_energy.cell(row=row + 1, column=1, value='Electricity Price')
        ws_energy.cell(row=row + 1, column=2, value=round(price_per_kwh, 5))
        ws_energy.cell(row=row + 1, column=3, value='$/kWh')
        row += 1
        
        ws_energy.cell(row=row + 1, column=1, value='Mining Efficiency')
        ws_energy.cell(row=row + 1, column=2, value=round(efficiency_w_th, 2))
        ws_energy.cell(row=row + 1, column=3, value='W/TH')
        row += 1
        
        ws_energy.cell(row=row + 1, column=1, value='Energy to Mining Ratio')
        ws_energy.cell(row=row + 1, column=2, value=round(energy_cost_percentage, 2))
        ws_energy.cell(row=row + 1, column=3, value='%')
        row += 1
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    platform_suffix = f'_{selected_platform_name.replace(" ", "_")}' if selected_platform else ''
    response['Content-Disposition'] = f'attachment; filename="forecasting_dashboard{platform_suffix}_export.xlsx"'
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(response)
    return response
