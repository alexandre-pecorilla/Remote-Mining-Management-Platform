from django.shortcuts import redirect
from django.contrib import messages
from openpyxl import load_workbook
from decimal import Decimal, InvalidOperation
from datetime import datetime
import logging
from ..models import RemoteMiningPlatform, Miner, Payout, Expense, TopUp

logger = logging.getLogger(__name__)


def import_platform_data(request):
    """Import platform data from uploaded Excel file"""
    if request.method == 'POST' and request.FILES.get('import_file'):
        try:
            file = request.FILES['import_file']
            wb = load_workbook(file)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(str(header).strip())

            # Process data rows
            imported_count = 0
            valid_fields = {f.name for f in RemoteMiningPlatform._meta.get_fields()}
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers):
                    cell_value = ws.cell(row=row, column=col_idx + 1).value
                    if cell_value is not None:
                        if isinstance(cell_value, datetime):
                            row_data[header] = cell_value.date()
                        elif isinstance(cell_value, str):
                            row_data[header] = cell_value.strip()
                        else:
                            row_data[header] = cell_value

                if row_data:
                    platform_data = {}
                    for field, value in row_data.items():
                        if field in valid_fields and value:
                            if field == 'energy_price':
                                platform_data[field] = Decimal(str(value))
                            else:
                                platform_data[field] = value

                    if platform_data:
                        RemoteMiningPlatform.objects.create(**platform_data)
                        imported_count += 1

            messages.success(request, f'Successfully imported {imported_count} platforms!')
            return redirect('platform_list')
            
        except Exception as e:
            messages.error(request, 'Wrong import file format or data. Please check your file and try again.')
            return redirect('platform_list')
    
    return redirect('platform_list')




def import_miner_data(request):
    """Import miner data from uploaded Excel file"""
    if request.method == 'POST' and request.FILES.get('import_file'):
        try:
            file = request.FILES['import_file']
            wb = load_workbook(file)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(str(header).strip())

            # Process data rows
            imported_count = 0
            valid_fields = {f.name for f in Miner._meta.get_fields()}
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers):
                    cell_value = ws.cell(row=row, column=col_idx + 1).value
                    if cell_value is not None:
                        if isinstance(cell_value, datetime):
                            row_data[header] = cell_value.date()
                        elif isinstance(cell_value, str):
                            row_data[header] = cell_value.strip()
                        else:
                            row_data[header] = cell_value

                if row_data:
                    miner_data = {}
                    for field, value in row_data.items():
                        if field in valid_fields and value:
                            if field == 'platform':
                                try:
                                    platform = RemoteMiningPlatform.objects.get(pk=int(float(value)))
                                    miner_data[field] = platform
                                except (RemoteMiningPlatform.DoesNotExist, ValueError, TypeError):
                                    continue
                            elif field in ['hashrate', 'power', 'efficiency', 'purchase_price']:
                                miner_data[field] = Decimal(str(value))
                            elif field in ['purchase_date', 'start_date']:
                                if isinstance(value, str):
                                    miner_data[field] = datetime.strptime(value, '%Y-%m-%d').date()
                                else:
                                    miner_data[field] = value
                            else:
                                miner_data[field] = value

                    if miner_data:
                        Miner.objects.create(**miner_data)
                        imported_count += 1

            messages.success(request, f'Successfully imported {imported_count} miners!')
            return redirect('miner_list')
            
        except Exception as e:
            messages.error(request, 'Wrong import file format or data. Please check your file and try again.')
            return redirect('miner_list')
    
    return redirect('miner_list')




def import_payout_data(request):
    """Import payout data from uploaded Excel file"""
    if request.method == 'POST' and request.FILES.get('import_file'):
        try:
            file = request.FILES['import_file']
            wb = load_workbook(file)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(str(header).strip())

            # Process data rows
            imported_count = 0
            valid_fields = {f.name for f in Payout._meta.get_fields()}
            for row in range(2, ws.max_row + 1):
                row_data = {}
                for col_idx, header in enumerate(headers):
                    cell_value = ws.cell(row=row, column=col_idx + 1).value
                    if cell_value is not None:
                        if isinstance(cell_value, datetime):
                            row_data[header] = cell_value.date()
                        elif isinstance(cell_value, str):
                            row_data[header] = cell_value.strip()
                        else:
                            row_data[header] = cell_value

                if row_data:
                    payout_data = {}
                    for field, value in row_data.items():
                        if field in valid_fields and value:
                            if field == 'platform':
                                try:
                                    platform = RemoteMiningPlatform.objects.get(pk=int(float(value)))
                                    payout_data[field] = platform
                                except (RemoteMiningPlatform.DoesNotExist, ValueError, TypeError):
                                    continue
                            elif field in ['payout_amount', 'closing_price']:
                                payout_data[field] = Decimal(str(value))
                            elif field == 'payout_date':
                                if isinstance(value, str):
                                    payout_data[field] = datetime.strptime(value, '%Y-%m-%d').date()
                                else:
                                    payout_data[field] = value
                            else:
                                payout_data[field] = value

                    if payout_data:
                        Payout.objects.create(**payout_data)
                        imported_count += 1

            messages.success(request, f'Successfully imported {imported_count} payouts!')
            return redirect('payout_list')
            
        except Exception as e:
            messages.error(request, 'Wrong import file format or data. Please check your file and try again.')
            return redirect('payout_list')
    
    return redirect('payout_list')


# Expense Import/Export Functions


def import_expense_data(request):
    """Import expense data from uploaded Excel file"""
    if request.method == 'POST' and request.FILES.get('import_file'):
        try:
            file = request.FILES['import_file']
            wb = load_workbook(file)
            ws = wb.active
            
            # Get headers from first row
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(row=1, column=col).value
                if header:
                    headers.append(str(header).lower().strip())

            # Process data rows
            imported_count = 0
            for row in range(2, ws.max_row + 1):
                expense_data = {}

                for col_idx, header in enumerate(headers):
                    cell_value = ws.cell(row=row, column=col_idx + 1).value

                    if header == 'expense_date' and cell_value:
                        try:
                            if isinstance(cell_value, datetime):
                                expense_data['expense_date'] = cell_value.date()
                            else:
                                expense_data['expense_date'] = datetime.strptime(str(cell_value), '%Y-%m-%d').date()
                        except (ValueError, TypeError) as e:
                            logger.warning("Expense import: bad date at row %d: %s", row, e)
                            continue
                    elif header == 'platform' and cell_value:
                        try:
                            platform_id = int(float(cell_value))
                            platform = RemoteMiningPlatform.objects.get(pk=platform_id)
                            expense_data['platform'] = platform
                        except (RemoteMiningPlatform.DoesNotExist, ValueError, TypeError):
                            pass
                    elif header == 'category' and cell_value:
                        category_value = str(cell_value).upper().strip()
                        if category_value in ['CAPEX', 'OPEX']:
                            expense_data['category'] = category_value
                    elif header == 'description' and cell_value:
                        expense_data['description'] = str(cell_value)
                    elif header == 'expense_amount' and cell_value:
                        try:
                            expense_data['expense_amount'] = Decimal(str(cell_value))
                        except (ValueError, InvalidOperation):
                            pass
                    elif header == 'invoice_link' and cell_value:
                        expense_data['invoice_link'] = str(cell_value)
                    elif header == 'receipt_link' and cell_value:
                        expense_data['receipt_link'] = str(cell_value)
                    elif header == 'notes' and cell_value:
                        expense_data['notes'] = str(cell_value)
                
                # Create expense if we have required fields
                if 'expense_date' in expense_data and 'category' in expense_data and 'expense_amount' in expense_data:
                    Expense.objects.create(**expense_data)
                    imported_count += 1
            
            messages.success(request, f'Successfully imported {imported_count} expenses!')
            return redirect('expense_list')
            
        except Exception as e:
            messages.error(request, 'Wrong import file format or data. Please check your file and try again.')
            return redirect('expense_list')
    
    return redirect('expense_list')


# ===== TOP-UP VIEWS =====



def import_topup_data(request):
    """Import top-up data from uploaded Excel file"""
    if request.method == 'POST' and request.FILES.get('import_file'):
        try:
            file = request.FILES['import_file']
            wb = load_workbook(file)
            ws = wb.active
            
            # Get headers from first row
            headers = [str(ws.cell(row=1, column=col).value).lower().strip() for col in range(1, ws.max_column + 1)]

            imported_count = 0

            # Process each row (skip header row)
            for row in range(2, ws.max_row + 1):
                topup_data = {}

                for col_idx, header in enumerate(headers):
                    cell_value = ws.cell(row=row, column=col_idx + 1).value

                    if header == 'topup_date' and cell_value:
                        try:
                            if isinstance(cell_value, datetime):
                                topup_data['topup_date'] = cell_value.date()
                            else:
                                topup_data['topup_date'] = datetime.strptime(str(cell_value), '%Y-%m-%d').date()
                        except (ValueError, TypeError) as e:
                            logger.warning("Top-up import: bad date at row %d: %s", row, e)
                            continue
                    elif header == 'platform' and cell_value:
                        try:
                            # Try to find platform by ID or name
                            if isinstance(cell_value, float):
                                platform = RemoteMiningPlatform.objects.get(id=int(cell_value))
                            else:
                                platform = RemoteMiningPlatform.objects.get(name=str(cell_value))
                            topup_data['platform'] = platform
                        except RemoteMiningPlatform.DoesNotExist:
                            continue
                    elif header == 'topup_amount' and cell_value:
                        try:
                            topup_data['topup_amount'] = float(cell_value)
                        except (ValueError, TypeError):
                            continue
                    elif header == 'description' and cell_value:
                        topup_data['description'] = str(cell_value)
                    elif header == 'receipt_link' and cell_value:
                        topup_data['receipt_link'] = str(cell_value)
                
                # Create top-up if we have required fields
                if 'topup_date' in topup_data and 'platform' in topup_data and 'topup_amount' in topup_data:
                    TopUp.objects.create(**topup_data)
                    imported_count += 1
            
            messages.success(request, f'Successfully imported {imported_count} top-ups!')
            return redirect('topup_list')
            
        except Exception as e:
            messages.error(request, 'Wrong import file format or data. Please check your file and try again.')
            return redirect('topup_list')
    
    return redirect('topup_list')


